# Controller/DeviceController.py

import json
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

from PySide6.QtCore import QObject, Signal, Slot, Property, QAbstractListModel, QModelIndex, Qt
from PySide6.QtCore import QUrl
from PySide6.QtQml import QJSValue
import pandas as pd

from DataManage.services.database_service import DatabaseService
from DataManage.models.device import DeviceType
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import logging
logger = logging.getLogger(__name__)


class DeviceListModel(QAbstractListModel):
    """设备列表模型"""

    # 定义角色
    IdRole = Qt.UserRole + 1
    TypeRole = Qt.UserRole + 2
    ManufacturerRole = Qt.UserRole + 3
    ModelRole = Qt.UserRole + 4
    SerialNumberRole = Qt.UserRole + 5
    StatusRole = Qt.UserRole + 6
    DescriptionRole = Qt.UserRole + 7
    CreatedAtRole = Qt.UserRole + 8
    DetailsRole = Qt.UserRole + 9

    def __init__(self, parent=None):
        super().__init__(parent)
        self._devices = []

    def rowCount(self, parent=QModelIndex()):
        return len(self._devices)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or index.row() >= len(self._devices):
            print('怎么他妈的进入这里了')
            return None

        device = self._devices[index.row()]
        # print('让我看啊看role到底是啥', role)
        # print("检查 role 字段",self.IdRole)
        # print("当前设备信息", device)
        # print("当前设备ID", device.get('id'))
        
        if role == self.IdRole:
            device_id = device.get('id')
            # print('yeah device_id:', device_id)
            return device_id
        elif role == self.TypeRole:
            return device.get('device_type', '')
        elif role == self.ManufacturerRole:
            return device.get('manufacturer', '')
        elif role == self.ModelRole:
            model_value = device.get('model', '')
            result = str(model_value) if model_value is not None else ''
            print(f'返回 model 值: "{result}" (原始值: {model_value})')
            return result  # 🔥 确保返回字符串
        elif role == self.SerialNumberRole:
            return device.get('serial_number', '')
        elif role == self.StatusRole:
            return device.get('status', 'active')
        elif role == self.DescriptionRole:
            return device.get('description', '')
        elif role == self.CreatedAtRole:
            return device.get('created_at', '')
        elif role == self.DetailsRole:
            # 返回设备详细信息的JSON字符串
            if device.get('device_type') == 'pump':
                return json.dumps(device.get('pump_details', {}))
            elif device.get('device_type') == 'motor':
                return json.dumps(device.get('motor_details', {}))
            elif device.get('device_type') == 'protector':
                return json.dumps(device.get('protector_details', {}))
            elif device.get('device_type') == 'separator':
                return json.dumps(device.get('separator_details', {}))
            return "{}"
        else:
            print("NOOOOOOOOOOOOOOOO, DeviceController ERROR, line77")

        return None

    def roleNames(self):
        return {
        self.IdRole: b'deviceId',
        self.TypeRole: b'deviceType',
        self.ManufacturerRole: b'manufacturer',
        self.ModelRole: b'deviceModel',
        self.SerialNumberRole: b'serialNumber',
        self.StatusRole: b'status',
        self.DescriptionRole: b'description',
        self.CreatedAtRole: b'createdAt',
        self.DetailsRole: b'details'
    }

    def setDevices(self, devices):
        """设置设备列表"""
        # print(f"这里是setDEVICE,设置设备列表: {len(devices)} 条记录")
        # 调试前3条数据
        # for i, dev in enumerate(devices[:3]):
        #     print(f"设备 {i+1}: ID={dev.get('id')}, 类型={dev.get('device_type')}")
    
        self.beginResetModel()
        self._devices = devices
        self.endResetModel()

    def getDevice(self, index):
        """获取指定索引的设备"""
        if 0 <= index < len(self._devices):
            return self._devices[index]
        return None

    def clear(self):
        """清空列表"""
        self.beginResetModel()
        self._devices = []
        self.endResetModel()


class DeviceController(QObject):
    """设备控制器"""

    # 信号定义
    deviceListChanged = Signal()
    deviceSaved = Signal(bool, str)  # success, message
    deviceDeleted = Signal(bool, str)  # success, message
    importCompleted = Signal(bool, str, int, int)  # success, message, successCount, errorCount
    exportCompleted = Signal(bool, str)  # success, filePath
    errorOccurred = Signal(str)  # errorMessage
    loadingChanged = Signal()
    selectedDeviceChanged = Signal()
    statisticsChanged = Signal()
    # 添加信号定义（在类的开头信号部分）
    pumpCurvesDataReady = Signal('QVariant')  # 泵性能曲线数据准备就绪
    # 🔥 新增导出相关信号
    exportCompleted = Signal(str, int)      # 文件路径, 导出数量
    exportProgress = Signal(int, int)       # 当前进度, 总数
    exportFailed = Signal(str)              # 错误信息
    # 🔥 添加模板生成相关信号
    templateGenerated = Signal(str)        # 模板生成成功
    templateGenerationFailed = Signal(str) # 模板生成失败

    def __init__(self, parent=None):
        super().__init__(parent)

        # 数据库服务
        self._db = DatabaseService()

        # 模型
        self._deviceListModel = DeviceListModel(self)

        # 状态
        self._loading = False
        self._currentPage = 1
        self._pageSize = 20
        self._totalCount = 0
        self._totalPages = 1
        self._currentFilter = {
            'device_type': None,
            'status': None,
            'keyword': ''
        }
        self._selectedDevice = None
        self._statistics = {
            'total_count': 0,
            'type_statistics': {},
            'status_statistics': {}
        }

        # 连接数据库信号
        self._db.deviceListUpdated.connect(self._onDeviceListUpdated)
        self._db.databaseError.connect(self._onDatabaseError)

        # 初始加载
        self.loadDevices()
        self.loadStatistics()

    # 属性
    @Property(QObject, notify=deviceListChanged)
    def deviceListModel(self):
        return self._deviceListModel

    @Property(bool, notify=loadingChanged)
    def loading(self):
        return self._loading

    @Property(int, notify=deviceListChanged)
    def currentPage(self):
        return self._currentPage

    @Property(int, notify=deviceListChanged)
    def totalPages(self):
        return self._totalPages

    @Property(int, notify=deviceListChanged)
    def totalCount(self):
        return self._totalCount

    @Property('QVariant', notify=selectedDeviceChanged)
    def selectedDevice(self):
        return self._selectedDevice or {}

    @Property('QVariant', notify=statisticsChanged)
    def statistics(self):
        return self._statistics

    def _setLoading(self, loading):
        if self._loading != loading:
            self._loading = loading
            self.loadingChanged.emit()

    def _onDeviceListUpdated(self):
        """设备列表更新时重新加载"""
        self.loadDevices()
        self.loadStatistics()

    def _onDatabaseError(self, error_msg):
        """处理数据库错误"""
        self.errorOccurred.emit(error_msg)

    # 设备列表操作
    @Slot()
    def loadDevices(self):
        """加载设备列表"""
        self._setLoading(True)

        try:
            # 调试信息
            # print(f"加载设备: 过滤条件={self._currentFilter}")
        
            # 如果有搜索关键词，使用搜索功能
            if self._currentFilter.get('keyword'):
                devices = self._db.search_devices(
                    keyword=self._currentFilter['keyword'],
                    device_type=self._currentFilter.get('device_type')
                )
                # 打印找到的设备
                # print(f"这里是loaddevice，检索到 {len(devices)} 个设备")
                # for device in devices[:3]:  # 只打印前三个
                #     print(f"设备: ID={device.get('id')}, 型号={device.get('modelDevice')}")
                
                self._deviceListModel.setDevices(devices)
                self._totalCount = len(devices)
                self._totalPages = 1
                self._currentPage = 1
            else:
                # 否则使用分页加载
                result = self._db.get_devices(
                    device_type=self._currentFilter.get('device_type'),
                    status=self._currentFilter.get('status'),
                    page=self._currentPage,
                    page_size=self._pageSize
                )

                self._deviceListModel.setDevices(result['devices'])
                self._totalCount = result['total_count']
                self._totalPages = result['total_pages']

            self.deviceListChanged.emit()

        except Exception as e:
            self.errorOccurred.emit(f"加载设备列表失败: {str(e)}")

        finally:
            self._setLoading(False)

    @Slot(str)
    def filterByType(self, device_type):
        """按类型筛选"""
        if device_type == "all":
            self._currentFilter['device_type'] = None
        else:
            self._currentFilter['device_type'] = device_type
        self._currentPage = 1
        self.loadDevices()

    @Slot(str)
    def filterByStatus(self, status):
        """按状态筛选"""
        if status == "all":
            self._currentFilter['status'] = None
        else:
            self._currentFilter['status'] = status
        self._currentPage = 1
        self.loadDevices()

    @Slot(str)
    def searchDevices(self, keyword):
        """搜索设备"""
        self._currentFilter['keyword'] = keyword.strip()
        self._currentPage = 1
        self.loadDevices()

    @Slot(int)
    def goToPage(self, page):
        """跳转到指定页"""
        if 1 <= page <= self._totalPages:
            self._currentPage = page
            self.loadDevices()

    @Slot()
    def nextPage(self):
        """下一页"""
        if self._currentPage < self._totalPages:
            self._currentPage += 1
            self.loadDevices()

    @Slot()
    def previousPage(self):
        """上一页"""
        if self._currentPage > 1:
            self._currentPage -= 1
            self.loadDevices()

    # 设备详情操作
    @Slot(int)
    def selectDevice(self, device_id):
        """选择设备查看详情"""
        self._setLoading(True)

        try:
            device = self._db.get_device_by_id(device_id)
            if device:
                self._selectedDevice = device
                self.selectedDeviceChanged.emit()
            else:
                self.errorOccurred.emit("设备不存在")

        except Exception as e:
            self.errorOccurred.emit(f"获取设备详情失败: {str(e)}")

        finally:
            self._setLoading(False)

    @Slot()
    def clearSelectedDevice(self):
        """清除选中的设备"""
        self._selectedDevice = None
        self.selectedDeviceChanged.emit()

    # 设备CRUD操作
    @Slot(str)
    def saveDevice(self, device_data):
        """保存设备（新建或更新）"""
        self._setLoading(True)


        try:
             # 解析JSON字符串为Python字典
            data = json.loads(device_data)
        
            # 转换QML传来的数据
            # data = dict(device_data)

            # 处理设备详情数据
            device_type = data.get('device_type')
            if device_type == 'pump':
                data['pump_details'] = data.get('pump_details', {})
            elif device_type == 'motor':
                data['motor_details'] = data.get('motor_details', {})
                # 处理频率参数
                freq_params = data['motor_details'].get('frequency_params', [])
                if isinstance(freq_params, str):
                    data['motor_details']['frequency_params'] = json.loads(freq_params)
            elif device_type == 'protector':
                data['protector_details'] = data.get('protector_details', {})
            elif device_type == 'separator':
                data['separator_details'] = data.get('separator_details', {})

            # 判断是新建还是更新
            device_id = data.get('id', -1)
            if device_id and device_id > 0:
                # 更新
                success = self._db.update_device(device_id, data)
                if success:
                    self.deviceSaved.emit(True, "设备更新成功")
                else:
                    self.deviceSaved.emit(False, "设备更新失败")
            else:
                # 新建
                new_id = self._db.create_device(data)
                if new_id:
                    self.deviceSaved.emit(True, "设备创建成功")
                else:
                    self.deviceSaved.emit(False, "设备创建失败")

        except Exception as e:
            self.deviceSaved.emit(False, str(e))
            self.errorOccurred.emit(f"保存设备失败: {str(e)}")

        finally:
            self._setLoading(False)

    @Slot(int)
    def deleteDevice(self, device_id):
        """删除设备"""
        self._setLoading(True)

        try:
            success = self._db.delete_device(device_id)
            if success:
                self.deviceDeleted.emit(True, "设备删除成功")
                # 如果删除的是当前选中的设备，清除选择
                if self._selectedDevice and self._selectedDevice.get('id') == device_id:
                    self.clearSelectedDevice()
            else:
                self.deviceDeleted.emit(False, "设备删除失败")

        except Exception as e:
            self.deviceDeleted.emit(False, str(e))
            self.errorOccurred.emit(f"删除设备失败: {str(e)}")

        finally:
            self._setLoading(False)

    @Slot(list)
    def batchDeleteDevices(self, device_ids):
        """批量删除设备"""
        self._setLoading(True)

        try:
            success = self._db.batch_delete_devices(device_ids)
            if success:
                self.deviceDeleted.emit(True, f"成功删除{len(device_ids)}个设备")
            else:
                self.deviceDeleted.emit(False, "批量删除失败")

        except Exception as e:
            self.deviceDeleted.emit(False, str(e))
            self.errorOccurred.emit(f"批量删除失败: {str(e)}")

        finally:
            self._setLoading(False)

    @Slot(str, str, bool)
    def importFromExcel(self, file_url, device_type, is_metric=False):
        """从Excel导入设备（支持双Sheet结构）"""
        self._setLoading(True)

        try:
            # 处理文件路径
            file_path = QUrl(file_url).toLocalFile()
            if not os.path.exists(file_path):
                raise ValueError("文件不存在")

            logger.info(f"开始导入设备: 文件={file_path}, 类型={device_type}, 公制={is_metric}")

            # 🔥 新增：检测Excel文件结构
            workbook_sheets = self._detect_excel_structure(file_path)
        
            if workbook_sheets.get('has_performance_sheet'):
                # 使用新的双Sheet结构处理
                result = self._import_with_performance_data(file_path, device_type, is_metric, workbook_sheets)
            else:
                # 使用原有的单Sheet结构处理
                result = self._import_legacy_format(file_path, device_type, is_metric)

            self.importCompleted.emit(
                True,
                f"导入完成：成功{result['success_count']}条，失败{result['error_count']}条",
                result['success_count'],
                result['error_count']
            )

            # 如果有错误，显示错误详情
            if result['errors']:
                error_details = "\n".join([
                    f"第{err['row']}行: {err['error']}"
                    for err in result['errors'][:5]  # 只显示前5个错误
                ])
                if len(result['errors']) > 5:
                    error_details += f"\n...还有{len(result['errors']) - 5}个错误"
                self.errorOccurred.emit(f"导入错误详情:\n{error_details}")

        except Exception as e:
            logger.error(f"导入失败: {str(e)}")
            self.importCompleted.emit(False, str(e), 0, 0)
            self.errorOccurred.emit(f"导入失败: {str(e)}")

        finally:
            self._setLoading(False)

    def _detect_excel_structure(self, file_path: str) -> Dict[str, Any]:
        """检测Excel文件结构"""
        try:
            # 获取所有工作表名称
            xl_file = pd.ExcelFile(file_path)
            sheet_names = xl_file.sheet_names
        
            structure = {
                'sheet_names': sheet_names,
                'has_performance_sheet': False,
                'basic_sheet_name': None,
                'performance_sheet_name': None
            }
        
            # 检测是否包含性能数据表
            performance_sheet_candidates = ['性能数据', 'Performance Data', 'Performance', '性能曲线']
            basic_sheet_candidates = ['基本设备信息', 'Basic Info', 'Device Info', '设备信息']
        
            for sheet in sheet_names:
                if sheet in performance_sheet_candidates:
                    structure['has_performance_sheet'] = True
                    structure['performance_sheet_name'] = sheet
                elif sheet in basic_sheet_candidates:
                    structure['basic_sheet_name'] = sheet
        
            # 如果没有明确的基本信息表，使用第一个表
            if not structure['basic_sheet_name'] and sheet_names:
                structure['basic_sheet_name'] = sheet_names[0]
        
            logger.info(f"检测到Excel结构: {structure}")
            return structure
        
        except Exception as e:
            logger.error(f"检测Excel结构失败: {str(e)}")
            return {'sheet_names': [], 'has_performance_sheet': False}

    def _import_with_performance_data(self, file_path: str, device_type: str, is_metric: bool, 
                                      structure: Dict[str, Any]) -> Dict[str, Any]:
        """使用新的双Sheet结构导入"""
        try:
            # 🔥 读取基本设备信息
            basic_sheet_name = structure['basic_sheet_name']
            basic_df = pd.read_excel(file_path, sheet_name=basic_sheet_name)
        
            # 🔥 读取性能数据
            performance_sheet_name = structure['performance_sheet_name']
            performance_df = pd.read_excel(file_path, sheet_name=performance_sheet_name)
        
            logger.info(f"基本信息表: {len(basic_df)} 行, 性能数据表: {len(performance_df)} 行")
        
            # 转换为记录格式
            basic_records = basic_df.to_dict('records')
            performance_records = performance_df.to_dict('records')
        
            # 🔥 按型号分组性能数据
            performance_by_model = {}
            for record in performance_records:
                model = str(record.get('型号', '')).strip()
                if model:
                    if model not in performance_by_model:
                        performance_by_model[model] = []
                    performance_by_model[model].append(record)
        
            logger.info(f"性能数据分组: {list(performance_by_model.keys())}")
        
            # 🔥 处理每个设备记录
            result = {
                'success_count': 0,
                'error_count': 0,
                'errors': []
            }
        
            for row_idx, basic_record in enumerate(basic_records, 2):  # 从第2行开始（第1行是表头）
                try:
                    # 获取设备型号
                    device_model = str(basic_record.get('型号', '')).strip()
                
                    if not device_model:
                        result['errors'].append({
                            'row': row_idx,
                            'error': '设备型号不能为空'
                        })
                        result['error_count'] += 1
                        continue
                
                    # 🔥 构建设备数据（包含基本信息和性能数据）
                    device_data = self._build_device_data_with_performance(
                        basic_record, 
                        performance_by_model.get(device_model, []), 
                        device_type, 
                        is_metric
                    )
                
                    # 🔥 调用数据库保存
                    device_id = self._save_device_to_database(device_data)
                
                    if device_id:
                        result['success_count'] += 1
                        logger.info(f"成功导入设备: {device_model} (ID: {device_id})")
                    else:
                        result['errors'].append({
                            'row': row_idx,
                            'error': f'保存设备失败: {device_model}'
                        })
                        result['error_count'] += 1
                    
                except Exception as e:
                    result['errors'].append({
                        'row': row_idx,
                        'error': f'处理第{row_idx}行时出错: {str(e)}'
                    })
                    result['error_count'] += 1
                    logger.error(f"处理第{row_idx}行失败: {str(e)}")
        
            return result
        
        except Exception as e:
            logger.error(f"双Sheet导入失败: {str(e)}")
            raise

    def _import_legacy_format(self, file_path: str, device_type: str, is_metric: bool) -> Dict[str, Any]:
        """使用原有的单Sheet结构导入"""
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            excel_data = df.to_dict('records')

            # 调用原有的数据库导入方法
            result = self._db.import_devices_from_(excel_data, device_type, is_metric)
        
            logger.info(f"单Sheet导入完成: 成功{result['success_count']}条")
            return result
        
        except Exception as e:
            logger.error(f"单Sheet导入失败: {str(e)}")
            raise

    def _build_device_data_with_performance(self, basic_record: Dict, performance_records: List[Dict], 
                                             device_type: str, is_metric: bool) -> Dict[str, Any]:
        """构建包含性能数据的设备信息"""
        try:
            # 🔥 构建基本设备数据
            device_data = {
                'device_type': device_type,
                'manufacturer': str(basic_record.get('制造商', '')),
                'model': str(basic_record.get('型号', '')),
                'series': str(basic_record.get('系列', '')),
                'serial_number': str(basic_record.get('序列号', '')),
                'status': str(basic_record.get('状态', 'active')),
                'description': str(basic_record.get('描述', '')),
            }
        
            # 🔥 根据设备类型处理详细信息
            if device_type == 'pump':
                device_data['pump_details'] = self._build_pump_details(basic_record, performance_records, is_metric)
            elif device_type == 'motor':
                device_data['motor_details'] = self._build_motor_details(basic_record, is_metric)
            elif device_type == 'protector':
                device_data['protector_details'] = self._build_protector_details(basic_record, is_metric)
            elif device_type == 'separator':
                device_data['separator_details'] = self._build_separator_details(basic_record, is_metric)
        
            return device_data
        
        except Exception as e:
            logger.error(f"构建设备数据失败: {str(e)}")
            raise

    def _build_pump_details(self, basic_record: Dict, performance_records: List[Dict], is_metric: bool) -> Dict[str, Any]:
        """构建泵的详细信息（包含性能数据）"""
        try:
            # 🔥 基本参数
            pump_details = {
                'impeller_model': str(basic_record.get('叶轮型号', '')),
                'displacement_min': self._parse_float(basic_record.get('最小流量(m³/d)' if is_metric else '最小流量(bbl/d)')),
                'displacement_max': self._parse_float(basic_record.get('最大流量(m³/d)' if is_metric else '最大流量(bbl/d)')),
                'single_stage_head': self._parse_float(basic_record.get('单级扬程(m)' if is_metric else '单级扬程(ft)')),
                'single_stage_power': self._parse_float(basic_record.get('单级功率(kW)' if is_metric else '单级功率(HP)')),
                'max_stages': self._parse_int(basic_record.get('最大级数')),
                'efficiency': self._parse_float(basic_record.get('效率(%)')),
                'outside_diameter': self._parse_float(basic_record.get('外径(mm)' if is_metric else '外径(in)')),
                'shaft_diameter': self._parse_float(basic_record.get('轴径(mm)' if is_metric else '轴径(in)')),
                'weight': self._parse_float(basic_record.get('重量(kg)' if is_metric else '重量(lbs)')),
                'mounting_height': self._parse_float(basic_record.get('装配高度(mm)' if is_metric else '长度(in)'))
            }
        
            # 🔥 处理性能曲线数据
            if performance_records:
                pump_details['performance_curves'] = self._process_pump_performance_data(performance_records, is_metric)
                logger.info(f"为泵 {basic_record.get('型号')} 添加了 {len(performance_records)} 个性能数据点")
        
            return pump_details
        
        except Exception as e:
            logger.error(f"构建泵详细信息失败: {str(e)}")
            raise

    def _process_pump_performance_data(self, performance_records: List[Dict], is_metric: bool) -> List[Dict]:
        """处理泵性能曲线数据"""
        curves = []
    
        try:
            for record in performance_records:
                curve_point = {
                    'frequency': self._parse_float(record.get('频率(Hz)', 60)),
                    'flow_rate': self._parse_float(record.get('流量(m³/d)' if is_metric else '流量(bbl/d)')),
                    'head': self._parse_float(record.get('扬程(m)' if is_metric else '扬程(ft)')),
                    'efficiency': self._parse_float(record.get('效率(%)')),
                    'power': self._parse_float(record.get('功率(kW)' if is_metric else '功率(HP)')),
                    'data_point_number': self._parse_int(record.get('数据点序号', 1)),
                    'data_source': str(record.get('数据来源', '导入数据')),
                    'test_date': str(record.get('测试日期', '')),
                    'notes': str(record.get('备注', ''))
                }
            
                # 🔥 验证数据完整性
                required_fields = ['flow_rate', 'head', 'efficiency', 'power']
                if all(curve_point[field] is not None for field in required_fields):
                    curves.append(curve_point)
                else:
                    logger.warning(f"性能数据点不完整，跳过: {curve_point}")
    
        except Exception as e:
            logger.error(f"处理性能曲线数据失败: {str(e)}")
    
        return curves

    def _save_device_to_database(self, device_data: Dict[str, Any]) -> Optional[int]:
        """保存设备到数据库"""
        try:
            # 🔥 检查是否为更新操作
            existing_device = None
            if device_data.get('serial_number'):
                existing_device = self._db.get_device_by_serial_number(device_data['serial_number'])
        
            if existing_device:
                # 更新现有设备
                success = self._db.update_device(existing_device['id'], device_data)
                return existing_device['id'] if success else None
            else:
                # 创建新设备
                new_id = self._db.create_device(device_data)
            
                # 🔥 如果有性能曲线数据，单独保存
                if device_data.get('pump_details', {}).get('performance_curves'):
                    self._save_performance_curves(new_id, device_data['pump_details']['performance_curves'])
            
                return new_id
    
        except Exception as e:
            logger.error(f"保存设备到数据库失败: {str(e)}")
            return None

    def _save_performance_curves(self, device_id: int, curves_data: List[Dict]):
        """保存性能曲线数据到数据库"""
        try:
            if hasattr(self._db, 'save_pump_curves'):
                # 🔥 如果数据库服务支持保存曲线数据
                model = self._db.get_device_by_id(device_id).get('model', '')
                curves_dict = {
                    'flow': [c['flow_rate'] for c in curves_data],
                    'head': [c['head'] for c in curves_data],
                    'efficiency': [c['efficiency'] for c in curves_data],
                    'power': [c['power'] for c in curves_data],
                    'standard_frequency': curves_data[0]['frequency'] if curves_data else 60,
                    'data_source': 'import_excel',
                    'version': '1.0'
                }
                self._db.save_pump_curves(model, curves_dict)
                logger.info(f"保存了设备 {device_id} 的性能曲线数据")
            else:
                logger.warning("数据库不支持保存性能曲线数据")
            
        except Exception as e:
            logger.error(f"保存性能曲线数据失败: {str(e)}")

    # 🔥 辅助方法
    def _parse_float(self, value, default=None):
        """安全地解析浮点数"""
        if value is None or value == '':
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default

    def _parse_int(self, value, default=None):
        """安全地解析整数"""
        if value is None or value == '':
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default



    @Slot(str, str, result=str)
    def exportToExcel(self, save_url, device_type):
        """导出设备到Excel"""
        self._setLoading(True)

        try:
            # 处理文件路径
            file_path = QUrl(save_url).toLocalFile()

            # 确定设备类型
            export_type = None if device_type == "all" else device_type

            # 获取数据
            data = self._db.export_devices_to_dict(export_type)

            if not data:
                self.exportCompleted.emit(False, "没有数据可导出")
                return file_path

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 根据设备类型设置列顺序
            if device_type == "pump":
                columns = [
                    'manufacturer', 'modelDevice', 'serial_number', 'status',
                    'impeller_model', 'displacement_min', 'displacement_max',
                    'single_stage_head', 'single_stage_power', 'shaft_diameter',
                    'mounting_height', 'outside_diameter', 'max_stages', 'efficiency',
                    'description', 'created_at'
                ]
            elif device_type == "motor":
                columns = [
                    'manufacturer', 'modelDevice', 'serial_number', 'status',
                    'motor_type', 'outside_diameter', 'length', 'weight',
                    'insulation_class', 'protection_class',
                    'power_50hz', 'voltage_50hz', 'current_50hz', 'speed_50hz',
                    'power_60hz', 'voltage_60hz', 'current_60hz', 'speed_60hz',
                    'description', 'created_at'
                ]
            elif device_type == "protector":
                columns = [
                    'manufacturer', 'modelDevice', 'serial_number', 'status',
                    'outer_diameter', 'length', 'weight', 'thrust_capacity',
                    'seal_type', 'max_temperature',
                    'description', 'created_at'
                ]
            elif device_type == "separator":
                columns = [
                    'manufacturer', 'modelDevice', 'serial_number', 'status',
                    'outer_diameter', 'length', 'weight',
                    'separation_efficiency', 'gas_handling_capacity',
                    'liquid_handling_capacity',
                    'description', 'created_at'
                ]
            else:
                # 所有设备类型，使用所有列
                columns = list(df.columns)

            # 重新排序列
            columns = [col for col in columns if col in df.columns]
            df = df[columns]

            # 保存到Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='设备数据', index=False)

                # 调整列宽
                worksheet = writer.sheets['设备数据']
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(str(col))
                    ) + 2
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 30)

            self.exportCompleted.emit(True, file_path)
            return file_path

        except Exception as e:
            self.exportCompleted.emit(False, "")
            self.errorOccurred.emit(f"导出失败: {str(e)}")
            return ""

        finally:
            self._setLoading(False)

    # 统计功能
    @Slot()
    def loadStatistics(self):
        """加载设备统计信息"""
        try:
            self._statistics = self._db.get_device_statistics()
            self.statisticsChanged.emit()

        except Exception as e:
            self.errorOccurred.emit(f"加载统计信息失败: {str(e)}")

    # 工具方法
    @Slot(result=list)
    def getDeviceTypes(self):
        """获取设备类型列表"""
        return [
            {'value': 'pump', 'label': '潜油离心泵', 'label_en': 'Centrifugal Pump'},
            {'value': 'motor', 'label': '电机', 'label_en': 'Motor'},
            {'value': 'protector', 'label': '保护器', 'label_en': 'Protector'},
            {'value': 'separator', 'label': '分离器', 'label_en': 'Separator'}
        ]

    @Slot(result=list)
    def getDeviceStatuses(self):
        """获取设备状态列表"""
        return [
            {'value': 'active', 'label': '正常', 'label_en': 'Active'},
            {'value': 'inactive', 'label': '停用', 'label_en': 'Inactive'},
            {'value': 'maintenance', 'label': '维护中', 'label_en': 'Maintenance'}
        ]

    @Slot(str, result=str)
    def getDeviceTypeLabel(self, device_type):
        """获取设备类型标签"""
        type_labels = {
            'pump': '潜油离心泵',
            'motor': '电机',
            'protector': '保护器',
            'separator': '分离器'
        }
        return type_labels.get(device_type, device_type)

    @Slot(str, result=str)
    def getStatusLabel(self, status):
        """获取状态标签"""
        status_labels = {
            'active': '正常',
            'inactive': '停用',
            'maintenance': '维护中'
        }
        return status_labels.get(status, status)

    @Slot(result=str)
    def getExcelTemplate(self, device_type):
        """获取Excel导入模板路径"""
        # 这里可以返回预定义的Excel模板文件路径
        template_dir = os.path.join(os.path.dirname(__file__), '..', 'templates')
        template_path = os.path.join(template_dir, f'{device_type}_template.xlsx')

        if os.path.exists(template_path):
            return QUrl.fromLocalFile(template_path).toString()
        else:
            return ""

    @Slot(str)
    def saveDeviceFromJson(self, formDataJson):
        """从JSON字符串保存设备"""
        self._setLoading(True)

        try:
            # 解析JSON字符串为Python字典
            data = json.loads(formDataJson)
        
            # 处理设备详情数据
            device_type = data.get('device_type')
            device_id = data.get('id')
        
            # 判断是新建还是更新
            if device_id and device_id > 0:
                # 更新
                success = self._db.update_device(device_id, data)
                if success:
                    self.deviceSaved.emit(True, "设备更新成功")
                else:
                    self.deviceSaved.emit(False, "设备更新失败")
            else:
                # 新建
                new_id = self._db.create_device(data)
                if new_id:
                    self.deviceSaved.emit(True, "设备创建成功")
                else:
                    self.deviceSaved.emit(False, "设备创建失败")

        except Exception as e:
            self.deviceSaved.emit(False, str(e))
            self.errorOccurred.emit(f"保存设备失败: {str(e)}")

        finally:
            self._setLoading(False)

    # 在DeviceController类中添加以下方法

    @Slot(int, result='QVariant')
    def getPumpPerformanceCurves(self, pump_id):
        """获取泵性能曲线数据"""
        try:
            # 获取泵的基本信息
            pump = self._db.get_device_by_id(pump_id)
            if not pump or pump.get('device_type') != 'pump':
                return {'has_data': False, 'error': 'pump_not_found'}
            
            pump_details = pump.get('pump_details', {})
        
            # 生成性能曲线数据（基于泵的参数）
            curves_data = self._generatePumpCurves(pump_details)
        
            return {
                'has_data': True,
                'pump_info': {
                    'manufacturer': pump.get('manufacturer', ''),
                    'model': pump.get('modelDevice', ''),
                    'impeller_model': pump_details.get('impeller_model', ''),
                    'stages': pump_details.get('max_stages', 87),
                    'outside_diameter': pump_details.get('outside_diameter', 5.62)
                },
                'baseCurves': curves_data['baseCurves'],
                'operatingPoints': curves_data['operatingPoints'],
                'performanceZones': curves_data['performanceZones']
            }
        
        except Exception as e:
            print(f"获取泵性能曲线失败: {str(e)}")
            return {'has_data': False, 'error': str(e)}

    @Slot(str, result='QVariant') 
    def getPumpCurvesByModel(self, pump_model):
        """根据泵型号获取性能曲线数据"""
        try:
            # 根据型号查找泵
            pumps = self._db.search_devices(keyword=pump_model, device_type='pump')
            if not pumps:
                return {'has_data': False, 'error': 'pump_model_not_found'}
            
            # 使用第一个匹配的泵
            pump = pumps[0]
            return self.getPumpPerformanceCurves(pump.get('id'))
        
        except Exception as e:
            print(f"根据型号获取泵性能曲线失败: {str(e)}")
            return {'has_data': False, 'error': str(e)}

    def _generatePumpCurves(self, pump_details):
        """生成泵性能曲线数据"""
        import numpy as np
    
        # 获取泵的基本参数
        single_stage_head = pump_details.get('single_stage_head', 12.0)  # ft
        single_stage_power = pump_details.get('single_stage_power', 2.5)  # HP
        efficiency = pump_details.get('efficiency', 75.0)  # %
        displacement_min = pump_details.get('displacement_min', 100)  # bbl/d
        displacement_max = pump_details.get('displacement_max', 2000)  # bbl/d
        max_stages = pump_details.get('max_stages', 87)
    
        # 生成流量范围（bbl/d）
        flow_points = np.linspace(displacement_min, displacement_max, 20)
    
        # 计算性能曲线（基于典型的离心泵特性）
        flow_normalized = flow_points / displacement_max
    
        # 扬程曲线（抛物线形状）
        head_curve = []
        for f_norm in flow_normalized:
            # 典型的扬程-流量关系：H = H0 * (1 - a*Q^2)
            head_coeff = 1.0 - 0.3 * (f_norm ** 2)
            head = single_stage_head * head_coeff
            head_curve.append(max(head, 0))
    
        # 效率曲线（钟形曲线）
        efficiency_curve = []
        for f_norm in flow_normalized:
            # 效率在60-80%流量范围内最高
            if f_norm < 0.3:
                eff = efficiency * (0.5 + 1.67 * f_norm)
            elif f_norm <= 0.8:
                eff = efficiency * (0.9 + 0.1 * np.cos(2 * np.pi * (f_norm - 0.55)))
            else:
                eff = efficiency * (1.1 - 0.6 * (f_norm - 0.8))
            efficiency_curve.append(max(min(eff, 100), 0))
    
        # 功率曲线（随流量增加）
        power_curve = []
        for i, f_norm in enumerate(flow_normalized):
            # 功率与流量和扬程相关
            flow_bbl_d = flow_points[i]
            head_ft = head_curve[i] 
            eff_percent = efficiency_curve[i]
        
            # 水力功率计算：P = ρ * g * Q * H / η
            flow_m3_s = flow_bbl_d * 0.158987 / 86400  # bbl/d to m³/s
            head_m = head_ft * 0.3048  # ft to m
            eff_decimal = eff_percent / 100
        
            if eff_decimal > 0.1:
                hydraulic_power = (1000 * 9.81 * flow_m3_s * head_m) / 1000  # kW
                shaft_power = hydraulic_power / eff_decimal
                power_hp = shaft_power * 1.341  # kW to HP
            else:
                power_hp = single_stage_power * f_norm
            
            power_curve.append(power_hp)
    
        # 关键工作点
        operating_points = [
            {
                'flow': displacement_max * 0.7,
                'head': single_stage_head * 0.85,
                'efficiency': efficiency,
                'power': single_stage_power * 0.9,
                'label': 'BEP'  # Best Efficiency Point
            }
        ]
    
        # 性能区域
        performance_zones = {
            'optimal': {
                'minFlow': displacement_max * 0.6,
                'maxFlow': displacement_max * 0.8,
                'label': '最佳效率区'
            },
            'acceptable': {
                'minFlow': displacement_max * 0.4,
                'maxFlow': displacement_max * 0.9,
                'label': '可接受区域'
            }
        }
    
        return {
            'baseCurves': {
                'flow': flow_points.tolist(),
                'head': head_curve,
                'efficiency': efficiency_curve,
                'power': power_curve
            },
            'operatingPoints': operating_points,
            'performanceZones': performance_zones
        }

    @Slot(str, result=str)
    def getPumpCurvesDataString(self, pump_model):
        """根据泵型号获取性能曲线数据（返回JSON字符串）"""
        try:
            import json
        
            # 根据型号查找泵
            if pump_model:
                pumps = self._db.search_devices(keyword=pump_model, device_type='pump')
                if pumps:
                    pump = pumps[0]
                    pump_details = pump.get('pump_details', {})
                
                    # 生成性能曲线数据
                    curves_data = self._generatePumpCurves(pump_details)
                
                    result = {
                        'has_data': True,
                        'pump_info': {
                            'manufacturer': pump.get('manufacturer', ''),
                            'model': pump.get('modelDevice', ''),
                            'impeller_model': pump_details.get('impeller_model', ''),
                            'stages': pump_details.get('max_stages', 87),
                            'outside_diameter': pump_details.get('outside_diameter', 5.62)
                        },
                        'baseCurves': curves_data['baseCurves'],
                        'operatingPoints': curves_data['operatingPoints'],
                        'performanceZones': curves_data['performanceZones']
                    }
                
                    return json.dumps(result, ensure_ascii=False)
        
            # 没有找到泵或型号为空，返回空数据
            return json.dumps({'has_data': False, 'error': 'pump_not_found'})
        
        except Exception as e:
            print(f"获取泵性能曲线失败: {str(e)}")
            return json.dumps({'has_data': False, 'error': str(e)})

    @Slot('QVariant', result=str) 
    def getPumpCurvesFromStepData(self, step_data):
        """从stepData中获取泵信息并返回性能曲线数据"""
        try:
            import json
            from PySide6.QtQml import QJSValue
        
            # 🔥 修复：处理QJSValue对象
            if isinstance(step_data, QJSValue):
                # 将QJSValue转换为Python对象
                step_data_dict = step_data.toVariant()
            else:
                step_data_dict = step_data
            
            print(f"转换后的stepData类型: {type(step_data_dict)}")
        
            # 从stepData中提取泵信息
            if not step_data_dict:
                return json.dumps({'has_data': False, 'error': 'no_step_data'})
            
            pump_info = step_data_dict.get('pump', {}) if hasattr(step_data_dict, 'get') else {}
        
            # 🔥 处理pump_info也可能是QJSValue的情况
            if isinstance(pump_info, QJSValue):
                pump_info = pump_info.toVariant()
        
            pump_model = pump_info.get('model', '') if pump_info else ''
            pump_id = pump_info.get('id', 0) if pump_info else 0
        
            print(f"从stepData获取泵信息: 型号={pump_model}, ID={pump_id}")
            print(f"泵信息详情: {pump_info}")
        
            # 如果有具体的泵参数，直接使用
            if pump_info and (pump_info.get('singleStageHead') or pump_info.get('stages')):
                curves_data = self._generatePumpCurvesFromParams(pump_info)
            
                result = {
                    'has_data': True,
                    'pump_info': {
                        'manufacturer': pump_info.get('manufacturer', 'Centrilift'),
                        'model': pump_model,
                        'stages': pump_info.get('stages', 87),
                        'outside_diameter': pump_info.get('outsideDiameter', 5.62)
                    },
                    'baseCurves': curves_data['baseCurves'],
                    'operatingPoints': curves_data['operatingPoints'],
                    'performanceZones': curves_data['performanceZones']
                }
            
                return json.dumps(result, ensure_ascii=False)
        
            # 否则尝试从数据库查找
            elif pump_model:
                return self.getPumpCurvesDataString(pump_model)
        
            # 都没有，返回默认数据
            else:
                curves_data = self._generateDefaultPumpCurves()
                result = {
                    'has_data': True,
                    'pump_info': {
                        'manufacturer': 'Centrilift',
                        'model': 'GN4000',
                        'stages': 87,
                        'outside_diameter': 5.62
                    },
                    'baseCurves': curves_data['baseCurves'],
                    'operatingPoints': curves_data['operatingPoints'],
                    'performanceZones': curves_data['performanceZones']
                }
            
                return json.dumps(result, ensure_ascii=False)
        
        except Exception as e:
            import traceback
            print(f"从stepData获取泵性能曲线失败: {str(e)}")
            print(f"错误堆栈: {traceback.format_exc()}")
            return json.dumps({'has_data': False, 'error': str(e)})

    @Slot(str, result=str)
    def getPumpCurvesFromStepDataString(self, step_data_json):
        """从stepData JSON字符串中获取泵信息并返回性能曲线数据"""
        try:
            import json
        
            # 解析JSON字符串
            step_data_dict = json.loads(step_data_json)
            print(f"解析stepData成功，keys: {list(step_data_dict.keys()) if step_data_dict else 'None'}")
        
            # 从stepData中提取泵信息
            if not step_data_dict:
                return json.dumps({'has_data': False, 'error': 'no_step_data'})
            
            pump_info = step_data_dict.get('pump', {})
            pump_model = pump_info.get('model', '')
            pump_id = pump_info.get('id', 0)
        
            print(f"从stepData获取泵信息: 型号={pump_model}, ID={pump_id}")
            print(f"泵信息详情: {pump_info}")
        
            # 如果有具体的泵参数，直接使用
            if pump_info and (pump_info.get('singleStageHead') or pump_info.get('stages')):
                curves_data = self._generatePumpCurvesFromParams(pump_info)
            
                result = {
                    'has_data': True,
                    'pump_info': {
                        'manufacturer': pump_info.get('manufacturer', 'Centrilift'),
                        'model': pump_model,
                        'stages': pump_info.get('stages', 87),
                        'outside_diameter': pump_info.get('outsideDiameter', 5.62)
                    },
                    'baseCurves': curves_data['baseCurves'],
                    'operatingPoints': curves_data['operatingPoints'],
                    'performanceZones': curves_data['performanceZones']
                }
            
                return json.dumps(result, ensure_ascii=False)
        
            # 否则尝试从数据库查找
            elif pump_model:
                return self.getPumpCurvesDataString(pump_model)
        
            # 都没有，返回默认数据
            else:
                print("使用默认泵参数生成性能曲线")
                curves_data = self._generateDefaultPumpCurves()
                result = {
                    'has_data': True,
                    'pump_info': {
                        'manufacturer': 'Centrilift',
                        'model': 'GN4000',
                        'stages': 87,
                        'outside_diameter': 5.62
                    },
                    'baseCurves': curves_data['baseCurves'],
                    'operatingPoints': curves_data['operatingPoints'],
                    'performanceZones': curves_data['performanceZones']
                }
            
                return json.dumps(result, ensure_ascii=False)
        
        except Exception as e:
            import traceback
            print(f"从stepData JSON获取泵性能曲线失败: {str(e)}")
            print(f"错误堆栈: {traceback.format_exc()}")
            return json.dumps({'has_data': False, 'error': str(e)})

    def _generatePumpCurvesFromParams(self, pump_info):
        """从泵参数生成性能曲线"""
        import numpy as np
    
        # 获取泵参数
        single_stage_head = pump_info.get('singleStageHead', 12.0)
        single_stage_power = pump_info.get('singleStagePower', 2.5)
        efficiency = pump_info.get('efficiency', 75.0)
        min_flow = pump_info.get('minFlow', 100)
        max_flow = pump_info.get('maxFlow', 2000)
        stages = pump_info.get('stages', 87)
    
        # 生成流量点
        flow_points = np.linspace(min_flow, max_flow, 25)
    
        # 计算多级泵的性能
        flow_normalized = flow_points / max_flow
    
        head_curve = []
        efficiency_curve = []
        power_curve = []
    
        for f_norm in flow_normalized:
            # 扬程曲线（多级）
            head_coeff = 1.0 - 0.25 * (f_norm ** 2)
            single_head = single_stage_head * head_coeff
            total_head = single_head * stages
            head_curve.append(max(total_head, 0))
        
            # 效率曲线
            if f_norm < 0.2:
                eff = efficiency * (0.4 + 3 * f_norm)
            elif f_norm <= 0.8:
                eff = efficiency * (0.85 + 0.15 * np.cos(np.pi * (f_norm - 0.5)))
            else:
                eff = efficiency * (1.0 - 0.4 * (f_norm - 0.8))
            efficiency_curve.append(max(min(eff, 95), 10))
        
            # 功率曲线（多级）
            power_factor = 0.3 + 0.7 * f_norm + 0.2 * (f_norm ** 2)
            total_power = single_stage_power * stages * power_factor
            power_curve.append(total_power)
    
        return {
            'baseCurves': {
                'flow': flow_points.tolist(),
                'head': head_curve,
                'efficiency': efficiency_curve,
                'power': power_curve
            },
            'operatingPoints': [
                {
                    'flow': max_flow * 0.7,
                    'head': single_stage_head * stages * 0.85,
                    'efficiency': efficiency,
                    'power': single_stage_power * stages * 0.8,
                    'label': 'BEP'
                }
            ],
            'performanceZones': {
                'optimal': {
                    'minFlow': max_flow * 0.6,
                    'maxFlow': max_flow * 0.8
                },
                'acceptable': {
                    'minFlow': max_flow * 0.4,
                    'maxFlow': max_flow * 0.9
                }
            }
        }


    def _generateDefaultPumpCurves(self):
        """生成默认泵性能曲线"""
        default_params = {
            'singleStageHead': 12.0,
            'singleStagePower': 2.5,
            'efficiency': 75.0,
            'minFlow': 100,
            'maxFlow': 2000,
            'stages': 87
        }
        return self._generatePumpCurvesFromParams(default_params)   

    @Slot(str, str)
    def exportDevices(self, file_url: str, device_type: str):
        """
        导出设备数据到Excel文件
        
        Args:
            file_url: 文件保存路径
            device_type: 设备类型 (all, pump, motor, protector, separator)
        """
        try:
            logger.info(f"开始导出设备数据: 类型={device_type}, 路径={file_url}")
            
            # 处理文件路径
            if file_url.startswith('file:///'):
                file_path = file_url[8:]  # Windows
            elif file_url.startswith('file://'):
                file_path = file_url[7:]  # Unix/Mac
            else:
                file_path = file_url
            
            file_path = os.path.normpath(file_path)
            
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 创建Excel工作簿
            workbook = openpyxl.Workbook()
            
            # 根据设备类型导出数据
            if device_type == "all":
                self._export_all_devices(workbook)
            else:
                self._export_single_device_type(workbook, device_type)
            
            # 保存文件
            workbook.save(file_path)
            
            # 获取导出的设备数量
            total_exported = self._count_exported_devices(device_type)
            
            logger.info(f"设备导出完成: {total_exported}个设备已导出到 {file_path}")
            self.exportCompleted.emit(file_path, total_exported)
            
        except Exception as e:
            error_msg = f"导出设备数据失败: {str(e)}"
            logger.error(error_msg)
            self.exportFailed.emit(error_msg)

    def _export_all_devices(self, workbook):
        """导出所有类型的设备"""
        device_types = ['pump', 'motor', 'protector', 'separator']
        
        # 移除默认工作表
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # 为每种设备类型创建工作表
        for device_type in device_types:
            self._export_single_device_type(workbook, device_type, create_new_sheet=True)

    def _export_single_device_type(self, workbook, device_type: str, create_new_sheet: bool = False):
        """导出单一类型的设备"""
        try:
            # 获取设备数据
            devices_data = self._db.get_devices(
                device_type=device_type.upper(),
                status='active'
            )
            
            devices = devices_data.get('devices', [])
            
            if not devices:
                logger.warning(f"没有找到{device_type}类型的设备数据")
                return
            
            # 创建或获取工作表
            if create_new_sheet:
                sheet_name = self._get_sheet_name(device_type)
                worksheet = workbook.create_sheet(sheet_name)
            else:
                worksheet = workbook.active
                worksheet.title = self._get_sheet_name(device_type)
            
            # 根据设备类型设置不同的列结构
            if device_type == 'pump':
                self._setup_pump_export_sheet(worksheet, devices)
            elif device_type == 'motor':
                self._setup_motor_export_sheet(worksheet, devices)
            elif device_type == 'protector':
                self._setup_protector_export_sheet(worksheet, devices)
            elif device_type == 'separator':
                self._setup_separator_export_sheet(worksheet, devices)
            
            logger.info(f"已导出{device_type}设备: {len(devices)}个")
            
        except Exception as e:
            logger.error(f"导出{device_type}设备失败: {e}")

    def _setup_pump_export_sheet(self, worksheet, devices):
        """设置泵设备导出表格"""
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 设置表头
        headers = [
            "设备ID", "制造商", "型号", "系列", "单级扬程(ft)", "单级功率(HP)",
            "最大级数", "最小流量(bbl/d)", "最大流量(bbl/d)", "效率(%)",
            "外径(in)", "轴径(in)", "重量(lbs)", "长度(in)", "状态", "创建时间", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 设置列宽
        column_widths = [10, 15, 15, 10, 12, 12, 10, 12, 12, 8, 10, 10, 10, 10, 8, 15, 20]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 填充数据
        for row, device in enumerate(devices, 2):
            pump_details = device.get('pump_details', {})
            
            data = [
                device.get('id', ''),
                device.get('manufacturer', ''),
                device.get('model', ''),
                device.get('series', ''),
                pump_details.get('single_stage_head', ''),
                pump_details.get('single_stage_power', ''),
                pump_details.get('max_stages', ''),
                pump_details.get('displacement_min', ''),
                pump_details.get('displacement_max', ''),
                pump_details.get('efficiency', ''),
                pump_details.get('outside_diameter', ''),
                pump_details.get('shaft_diameter', ''),
                pump_details.get('weight', ''),
                pump_details.get('length', ''),
                device.get('status', ''),
                device.get('created_at', ''),
                device.get('description', '')
            ]
            
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _setup_motor_export_sheet(self, worksheet, devices):
        print("1552,正在导出电机设备...")
        """设置电机导出表格（包含电流电压信息）"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 🔥 修改表头，添加电流电压字段
        headers = [
            "设备ID", "制造商", "型号", "系列", "序列号", "状态", "描述",
            "电机类型", "外径(mm)", "长度(mm)", "重量(kg)", 
            "绝缘等级", "防护等级",
            # 🔥 新增：电流电压字段
            "50Hz功率(kW)", "50Hz电压(V)", "50Hz电流(A)", "50Hz转速(rpm)",
            "60Hz功率(kW)", "60Hz电压(V)", "60Hz电流(A)", "60Hz转速(rpm)",
            "创建时间"
        ]

        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 设置列宽
        column_widths = [8, 15, 20, 10, 15, 8, 25, 15, 12, 12, 12, 12, 12, 
                         12, 12, 12, 12, 12, 12, 12, 12, 15]  # 🔥 增加新列的宽度
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 🔥 修改数据填充，添加频率参数处理
        for row, device in enumerate(devices, 2):
            motor_details = device.get('motor_details', {})
            frequency_params = motor_details.get('frequency_params', [])
        
            # 🔥 提取不同频率的参数
            freq_50_data = next((p for p in frequency_params if p.get('frequency') == 50), {})
            freq_60_data = next((p for p in frequency_params if p.get('frequency') == 60), {})

            data = [
                device.get('id', ''),
                device.get('manufacturer', ''),
                device.get('model', ''),
                device.get('series', ''),
                device.get('serial_number', ''),
                device.get('status', ''),
                device.get('description', ''),
                motor_details.get('motor_type', ''),
                motor_details.get('outside_diameter', ''),
                motor_details.get('length', ''),
                motor_details.get('weight', ''),
                motor_details.get('insulation_class', ''),
                motor_details.get('protection_class', ''),
                # 🔥 50Hz参数
                freq_50_data.get('power', ''),
                freq_50_data.get('voltage', ''),
                freq_50_data.get('current', ''),
                freq_50_data.get('speed', ''),
                # 🔥 60Hz参数
                freq_60_data.get('power', ''),
                freq_60_data.get('voltage', ''),
                freq_60_data.get('current', ''),
                freq_60_data.get('speed', ''),
                device.get('created_at', '')
            ]

            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _setup_protector_export_sheet(self, worksheet, devices):
        """设置保护器导出表格（基于实际数据库结构）"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
        # 🔥 根据实际数据库字段设置表头
        headers = [
            "设备ID", "制造商", "型号", "系列", "序列号", "状态", "描述",
            "外径(mm)", "长度(mm)", "重量(kg)", 
            "推力承载能力(kN)", "密封类型", "最高温度(°C)",
            "创建时间"
        ]
    
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
        # 设置列宽
        column_widths = [8, 15, 20, 10, 15, 8, 25, 12, 12, 12, 15, 12, 12, 15]
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
        # 填充数据
        for row, device in enumerate(devices, 2):
            protector_details = device.get('protector_details', {})
        
            data = [
                device.get('id', ''),
                device.get('manufacturer', ''),
                device.get('model', ''),
                device.get('series', ''),
                device.get('serial_number', ''),
                device.get('status', ''),
                device.get('description', ''),
                protector_details.get('outer_diameter', ''),
                protector_details.get('length', ''),
                protector_details.get('weight', ''),
                protector_details.get('thrust_capacity', ''),
                protector_details.get('seal_type', ''),
                protector_details.get('max_temperature', ''),
                device.get('created_at', '')
            ]
        
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _setup_separator_export_sheet(self, worksheet, devices):
        """设置分离器导出表格（基于实际数据库结构）"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
        # 🔥 根据实际数据库字段设置表头
        headers = [
            "设备ID", "制造商", "型号", "系列", "序列号", "状态", "描述",
            "外径(mm)", "长度(mm)", "重量(kg)", 
            "分离效率(%)", "气体处理能力(m³/d)", "液体处理能力(m³/d)",
            "创建时间"
        ]
    
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
        # 设置列宽
        column_widths = [8, 15, 20, 10, 15, 8, 25, 12, 12, 12, 12, 18, 18, 15]
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
        # 填充数据
        for row, device in enumerate(devices, 2):
            separator_details = device.get('separator_details', {})
        
            data = [
                device.get('id', ''),
                device.get('manufacturer', ''),
                device.get('model', ''),
                device.get('series', ''),
                device.get('serial_number', ''),
                device.get('status', ''),
                device.get('description', ''),
                separator_details.get('outer_diameter', ''),
                separator_details.get('length', ''),
                separator_details.get('weight', ''),
                separator_details.get('separation_efficiency', ''),
                separator_details.get('gas_handling_capacity', ''),
                separator_details.get('liquid_handling_capacity', ''),
                device.get('created_at', '')
            ]
        
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _get_sheet_name(self, device_type: str) -> str:
        """获取工作表名称"""
        names = {
            'pump': '潜油离心泵',
            'motor': '电机',
            'protector': '保护器',
            'separator': '分离器'
        }
        return names.get(device_type, device_type)

    def _count_exported_devices(self, device_type: str) -> int:
        """统计导出的设备数量"""
        try:
            if device_type == "all":
                total = 0
                for dtype in ['pump', 'motor', 'protector', 'separator']:
                    devices_data = self._db.get_devices(
                        device_type=dtype.upper(),
                        status='active'
                    )
                    total += len(devices_data.get('devices', []))
                return total
            else:
                devices_data = self._db.get_devices(
                    device_type=device_type.upper(),
                    status='active'
                )
                return len(devices_data.get('devices', []))
        except:
            return 0

    @Slot(str, str, bool)
    def generateTemplate(self, device_type: str, save_path: str, is_metric: bool):
        """
        生成设备导入模板
    
        Args:
            device_type: 设备类型 (pump, motor, protector, separator)
            save_path: 保存路径
            is_metric: 是否使用公制单位
        """
        try:
            logger.info(f"生成{device_type}模板: 单位制={is_metric}, 保存到={save_path}")
        
            # 处理文件路径
            if save_path.startswith('file:///'):
                file_path = save_path[8:]  # Windows
            elif save_path.startswith('file://'):
                file_path = save_path[7:]  # Unix/Mac
            else:
                file_path = save_path
        
            file_path = os.path.normpath(file_path)
        
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
            # 创建Excel工作簿
            workbook = openpyxl.Workbook()
        
            # 根据设备类型生成不同模板
            if device_type == 'pump':
                self._generate_pump_template(workbook, is_metric)
            elif device_type == 'motor':
                self._generate_motor_template(workbook, is_metric)
            elif device_type == 'protector':
                self._generate_protector_template(workbook, is_metric)
            elif device_type == 'separator':
                self._generate_separator_template(workbook, is_metric)
            else:
                raise ValueError(f"不支持的设备类型: {device_type}")
        
            # 保存文件
            workbook.save(file_path)
        
            logger.info(f"模板生成成功: {file_path}")
            self.templateGenerated.emit(file_path)
        
        except Exception as e:
            error_msg = f"生成模板失败: {str(e)}"
            logger.error(error_msg)
            self.templateGenerationFailed.emit(error_msg)

    def _generate_pump_template(self, workbook, is_metric: bool):
        """生成改进的泵设备导入模板（包含性能数据工作表）"""
    
        # ========== Sheet1: 基本设备信息 ==========
        if 'Sheet' in workbook.sheetnames:
            basic_sheet = workbook['Sheet']
        else:
            basic_sheet = workbook.create_sheet()
    
        basic_sheet.title = "基本设备信息"
    
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
        # 🔥 简化的基本信息表头（移除性能曲线数据）
        if is_metric:
            basic_headers = [
                "制造商", "型号", "系列", "举升方式", "序列号", "状态", "描述",
                "叶轮型号", "最小流量(m³/d)", "最大流量(m³/d)", 
                "单级扬程(m)", "单级功率(kW)", "最大级数", "效率(%)",
                "外径(mm)", "轴径(mm)", "重量(kg)", "装配高度(mm)",
                # # 应用范围
                # "最低温度(°C)", "最高温度(°C)", "最低压力(MPa)", "最高压力(MPa)", 
                # "最低粘度(mPa·s)", "最高粘度(mPa·s)"
            ]
        
            basic_example = [
                "Baker Hughes", "FLEXPump™ 400", "400", "esp", "BH-ESP-400-001", "active", "高效ESP泵",
                "D400", "50", "1500", "7.6", "1.9", "400", "68",
                "101.6", "19.05", "45.4", "1219",
                # "4", "121", "0.1", "13.8", "0.5", "1000"
            ]
        else:
            basic_headers = [
                "制造商", "型号", "系列", "举升方式", "序列号", "状态", "描述",
                "叶轮型号", "最小流量(bbl/d)", "最大流量(bbl/d)",
                "单级扬程(ft)", "单级功率(HP)", "最大级数", "效率(%)",
                "外径(in)", "轴径(in)", "重量(lbs)", "长度(in)",
                # "最低温度(°F)", "最高温度(°F)", "最低压力(psi)", "最高压力(psi)", 
                # "最低粘度(cp)", "最高粘度(cp)"
            ]
        
            basic_example = [
                "Baker Hughes", "FLEXPump™ 400", "400", "esp", "BH-ESP-400-001", "active", "高效ESP泵",
                "D400", "315", "9450", "25", "2.5", "400", "68",
                "4.0", "0.75", "100", "48",
                # "40", "250", "15", "2000", "0.5", "1000"
            ]
    
        # 设置基本信息表头
        for col, header in enumerate(basic_headers, 1):
            cell = basic_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
        # 填充基本信息示例数据
        for col, value in enumerate(basic_example, 1):
            cell = basic_sheet.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
        # 设置列宽
        for col in range(1, len(basic_headers) + 1):
            basic_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
        # ========== Sheet2: 性能数据 ==========
        performance_sheet = workbook.create_sheet("性能数据")
    
        # 🔥 性能数据表头设计
        if is_metric:
            perf_headers = [
                "型号",  # 🔥 关键：用于与Sheet1对齐
                "频率(Hz)", 
                "流量(m³/d)", "扬程(m)", "效率(%)", "功率(kW)",
                "数据点序号", "数据来源", "测试日期", "备注"
            ]
        
            # 🔥 示例性能数据（多个数据点）
            perf_examples = [
                ["FLEXPump™ 400", 60, 200, 25.0, 45, 3.2, 1, "厂商测试", "2024-01-15", "低流量点"],
                ["FLEXPump™ 400", 60, 400, 23.0, 58, 5.8, 2, "厂商测试", "2024-01-15", ""],
                ["FLEXPump™ 400", 60, 600, 21.0, 68, 8.1, 3, "厂商测试", "2024-01-15", "BEP点"],
                ["FLEXPump™ 400", 60, 800, 18.0, 65, 10.2, 4, "厂商测试", "2024-01-15", ""],
                ["FLEXPump™ 400", 60, 1000, 15.0, 58, 12.1, 5, "厂商测试", "2024-01-15", "高流量点"],
                ["FLEXPump™ 400", 50, 167, 20.8, 42, 2.8, 1, "厂商测试", "2024-01-15", "50Hz低流量"],
                ["FLEXPump™ 400", 50, 333, 19.2, 55, 5.1, 2, "厂商测试", "2024-01-15", "50Hz BEP"],
                ["FLEXPump™ 400", 50, 500, 17.5, 65, 7.2, 3, "厂商测试", "2024-01-15", "50Hz高效点"],
            ]
        else:
            perf_headers = [
                "型号",  # 🔥 关键：用于与Sheet1对齐
                "频率(Hz)", 
                "流量(bbl/d)", "扬程(ft)", "效率(%)", "功率(HP)",
                "数据点序号", "数据来源", "测试日期", "备注"
            ]
        
            perf_examples = [
                ["FLEXPump™ 400", 60, 1260, 82, 45, 4.3, 1, "厂商测试", "2024-01-15", "低流量点"],
                ["FLEXPump™ 400", 60, 2520, 75, 58, 7.8, 2, "厂商测试", "2024-01-15", ""],
                ["FLEXPump™ 400", 60, 3780, 69, 68, 10.9, 3, "厂商测试", "2024-01-15", "BEP点"],
                ["FLEXPump™ 400", 60, 5040, 59, 65, 13.7, 4, "厂商测试", "2024-01-15", ""],
                ["FLEXPump™ 400", 60, 6300, 49, 58, 16.2, 5, "厂商测试", "2024-01-15", "高流量点"],
                ["FLEXPump™ 400", 50, 1050, 68, 42, 3.7, 1, "厂商测试", "2024-01-15", "50Hz低流量"],
                ["FLEXPump™ 400", 50, 2100, 63, 55, 6.8, 2, "厂商测试", "2024-01-15", "50Hz BEP"],
                ["FLEXPump™ 400", 50, 3150, 57, 65, 9.6, 3, "厂商测试", "2024-01-15", "50Hz高效点"],
            ]
    
        # 设置性能数据表头样式
        perf_header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    
        for col, header in enumerate(perf_headers, 1):
            cell = performance_sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = perf_header_fill
            cell.alignment = header_alignment
            cell.border = border
    
        # 填充性能数据示例
        for row_idx, example_row in enumerate(perf_examples, 2):
            for col, value in enumerate(example_row, 1):
                cell = performance_sheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
    
        # 设置性能数据列宽
        perf_column_widths = [20, 10, 12, 12, 10, 12, 12, 15, 12, 20]
        for col, width in enumerate(perf_column_widths, 1):
            performance_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
        # ========== Sheet3: 使用说明 ==========
        instruction_sheet = workbook.create_sheet("使用说明")
        self._add_enhanced_template_instructions(instruction_sheet, is_metric)
    
        # 🔥 添加数据验证
        self._add_data_validation(basic_sheet, performance_sheet, is_metric)

    def _add_enhanced_template_instructions(self, worksheet, is_metric: bool):
        """添加增强的模板使用说明"""
        instructions = [
            "📋 泵设备导入模板使用说明",
            "",
            "🎯 模板结构：",
            "   Sheet1 - 基本设备信息：设备的基本参数和规格",
            "   Sheet2 - 性能数据：详细的性能曲线数据点",
            "   Sheet3 - 使用说明：本说明文档",
            "",
            "📊 Sheet1 (基本设备信息) 填写指南：",
            "   1. 制造商：设备制造商名称 (如：Baker Hughes, Schlumberger)",
            "   2. 型号：完整设备型号 (如：FLEXPump™ 400) - ⚠️ 必须与Sheet2中的型号完全一致",
            "   3. 系列：产品系列代码 (如：400, 500, 600)",
            "   4. 举升方式：esp/pcp/jet/espcp/hpp",
            "   5. 序列号：设备唯一序列号",
            "   6. 状态：active/inactive/maintenance",
            "",
            "🚀 Sheet2 (性能数据) 填写指南：",
            "   🔑 关键说明：型号列必须与Sheet1中的型号完全匹配！",
            "   ",
            "   列说明：",
            "   - 型号：与Sheet1对应的设备型号",
            "   - 频率：工作频率 (通常50Hz或60Hz)",
            f"   - 流量：{'m³/d' if is_metric else 'bbl/d'}",
            f"   - 扬程：{'m' if is_metric else 'ft'}",
            "   - 效率：百分比 (%)",
            f"   - 功率：{'kW' if is_metric else 'HP'}",
            "   - 数据点序号：同一型号同一频率下的点序号",
            "   - 数据来源：厂商测试/现场测试/计算值",
            "   - 测试日期：数据获取日期",
            "   - 备注：补充说明 (如：BEP点、低流量点等)",
            "",
            "💡 数据录入技巧：",
            "   1. 复制粘贴支持：",
            "      - 可以直接从厂商技术手册复制数据",
            "      - 支持从其他Excel文件批量粘贴",
            "      - 可以按列粘贴流量、扬程、效率、功率数据",
            "",
            "   2. 多频率支持：",
            "      - 同一型号可以有多个频率的数据",
            "      - 每个频率建议至少5-10个数据点",
            "      - 数据点应覆盖从关断到最大流量",
            "",
            "   3. 数据质量要求：",
            "      - 流量点应按升序排列",
            "      - 扬程数据应符合离心泵特性",
            "      - 效率曲线应呈现钟形分布",
            "      - 功率随流量递增",
            "",
            "⚠️ 常见错误避免：",
            "   1. 型号不匹配：Sheet1和Sheet2中的型号必须完全一致",
            "   2. 单位混用：确保所有数据使用统一单位制",
            "   3. 数据缺失：关键参数不能为空",
            "   4. 数值异常：检查是否存在明显不合理的数值",
            "",
            "🔄 导入流程：",
            "   1. 填写Sheet1的基本设备信息",
            "   2. 在Sheet2中录入对应型号的性能数据",
            "   3. 确保型号匹配",
            "   4. 保存文件并导入系统",
            "   5. 系统会自动关联并生成完整的性能曲线",
            "",
        ]
    
        # 设置说明样式
        title_font = Font(bold=True, size=16, color="1F4E79")
        section_font = Font(bold=True, size=12, color="2E75B6")
        normal_font = Font(size=10)
    
        for i, instruction in enumerate(instructions, 1):
            cell = worksheet.cell(row=i, column=1, value=instruction)
        
            if instruction.startswith("📋"):
                cell.font = title_font
            elif instruction.startswith(("🎯", "📊", "🚀", "💡", "⚠️", "🔄", "📞")):
                cell.font = section_font
            else:
                cell.font = normal_font
        
            # 合并单元格以便长文本显示
            if instruction.strip():
                worksheet.merge_cells(f"A{i}:H{i}")
    
        # 设置列宽
        worksheet.column_dimensions['A'].width = 80

    def _add_data_validation(self, basic_sheet, performance_sheet, is_metric: bool):
        """为模板添加数据验证"""
        from openpyxl.worksheet.datavalidation import DataValidation
    
        # 🔥 基本信息表验证
        # 举升方式验证
        lift_method_validation = DataValidation(
            type="list",
            formula1='"esp,pcp,jet,espcp,hpp"',
            showErrorMessage=True,
            errorTitle="无效的举升方式",
            error="请选择：esp, pcp, jet, espcp, hpp"
        )
        basic_sheet.add_data_validation(lift_method_validation)
        lift_method_validation.add("D2:D100")  # 举升方式列
    
        # 状态验证
        status_validation = DataValidation(
            type="list", 
            formula1='"active,inactive,maintenance"',
            showErrorMessage=True,
            errorTitle="无效的状态",
            error="请选择：active, inactive, maintenance"
        )
        basic_sheet.add_data_validation(status_validation)
        status_validation.add("F2:F100")  # 状态列
    
        # 🔥 性能数据表验证
        # 频率验证
        frequency_validation = DataValidation(
            type="list",
            formula1='"50,60"',
            showErrorMessage=True,
            errorTitle="无效的频率",
            error="请选择：50 或 60 Hz"
        )
        performance_sheet.add_data_validation(frequency_validation)
        frequency_validation.add("B2:B1000")  # 频率列
    
        # 效率范围验证 (0-100%)
        efficiency_validation = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=100,
            showErrorMessage=True,
            errorTitle="效率范围错误", 
            error="效率应在0-100%之间"
        )
        performance_sheet.add_data_validation(efficiency_validation)
        efficiency_validation.add("E2:E1000")  # 效率列
    
        # 数据来源验证
        source_validation = DataValidation(
            type="list",
            formula1='"厂商测试,现场测试,计算值,估算值"',
            showErrorMessage=True,
            errorTitle="无效的数据来源",
            error="请选择：厂商测试, 现场测试, 计算值, 估算值"
        )
        performance_sheet.add_data_validation(source_validation)
        source_validation.add("H2:H1000")  # 数据来源列

    def _generate_motor_template(self, workbook, is_metric: bool):
        """生成电机导入模板（修复版 - 包含电流电压等频率参数）"""
        # 设置工作表
        if 'Sheet' in workbook.sheetnames:
            worksheet = workbook['Sheet']
        else:
            worksheet = workbook.create_sheet()

        worksheet.title = "电机导入模板"

        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 🔥 修复：添加完整的表头（包含频率参数字段）
        if is_metric:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（公制）
                "电机类型", "外径(mm)", "长度(mm)", "重量(kg)",
                "绝缘等级", "防护等级",
                # 🔥 新增：50Hz频率参数
                "50Hz功率(kW)", "50Hz电压(V)", "50Hz电流(A)", "50Hz转速(rpm)",
                # 🔥 新增：60Hz频率参数
                "60Hz功率(kW)", "60Hz电压(V)", "60Hz电流(A)", "60Hz转速(rpm)"
            ]
        
            example_data = [
                "Baker Hughes", "Electrospeed 3", "ES3", "BH-ES3-001", "active", "高效潜油电机",
                "三相感应电机", "114.3", "3048", "136.1", "F", "IP68",
                # 50Hz参数
                "153", "3300", "28.5", "2950",
                # 60Hz参数
                "184", "3300", "34.2", "3540"
            ]
        else:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（英制）
                "电机类型", "外径(in)", "长度(in)", "重量(lbs)",
                "绝缘等级", "防护等级",
                # 🔥 新增：50Hz频率参数
                "50Hz功率(HP)", "50Hz电压(V)", "50Hz电流(A)", "50Hz转速(rpm)",
                # 🔥 新增：60Hz频率参数
                "60Hz功率(HP)", "60Hz电压(V)", "60Hz电流(A)", "60Hz转速(rpm)"
            ]
        
            example_data = [
                "Baker Hughes", "Electrospeed 3", "ES3", "BH-ES3-001", "active", "高效潜油电机",
                "三相感应电机", "4.5", "120", "300", "F", "IP68",
                # 50Hz参数
                "205", "3300", "38.2", "2950",
                # 60Hz参数
                "246", "3300", "45.8", "3540"
            ]

        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 🔥 调整列宽以适应新增的字段
        column_widths = [15, 20, 10, 15, 8, 25, 15, 12, 12, 12, 12, 12,
                         12, 12, 12, 12, 12, 12, 12, 12]
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 填充示例数据
        for col, value in enumerate(example_data, 1):
            cell = worksheet.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # 🔥 添加更多包含频率参数的示例数据
        additional_examples = []
        if is_metric:
            additional_examples = [
                ["Schlumberger", "MaxForce Motor", "MF450", "SLB-MF450-001", "active", "高功率潜油电机",
                 "永磁同步电机", "101.6", "2743", "95.3", "H", "IP67",
                 # 50Hz参数
                 "125", "4000", "20.8", "2950",
                 # 60Hz参数
                 "150", "4000", "25.0", "3540"],
             
                ["Weatherford", "RedaMax Motor", "RM300", "WFT-RM300-001", "active", "标准潜油电机",
                 "三相感应电机", "88.9", "2134", "68.0", "F", "IP55",
                 # 50Hz参数
                 "94", "2300", "26.4", "2950",
                 # 60Hz参数
                 "113", "2300", "31.7", "3540"]
            ]
        else:
            additional_examples = [
                ["Schlumberger", "MaxForce Motor", "MF450", "SLB-MF450-001", "active", "高功率潜油电机",
                 "永磁同步电机", "4.0", "108", "210", "H", "IP67",
                 # 50Hz参数
                 "167", "4000", "27.8", "2950",
                 # 60Hz参数
                 "200", "4000", "33.3", "3540"],
             
                ["Weatherford", "RedaMax Motor", "RM300", "WFT-RM300-001", "active", "标准潜油电机",
                 "三相感应电机", "3.5", "84", "150", "F", "IP55",
                 # 50Hz参数
                 "125", "2300", "35.2", "2950",
                 # 60Hz参数
                 "150", "2300", "42.3", "3540"]
            ]

        # 添加额外示例行
        for row_idx, example in enumerate(additional_examples, 3):
            for col, value in enumerate(example, 1):
                cell = worksheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 🔥 添加详细的使用说明
        self._add_motor_template_notes_with_frequency(worksheet, is_metric, len(headers))

    def _add_motor_template_notes_with_frequency(self, worksheet, is_metric: bool, header_count: int):
        """添加包含频率参数说明的电机模板说明"""
        notes_start_row = 6
    
        notes = [
            "📋 电机导入模板使用说明（包含频率参数）：",
            "",
            "🔧 数据库字段对应：",
            "   本模板设计用于导入电机基本信息和频率参数",
            "   数据将自动分别存储到 device_motors 和 motor_frequency_params 表",
            "",
            "1. 基本信息字段：",
            "   - 制造商：设备制造商名称",
            "   - 型号：完整设备型号",
            "   - 系列：产品系列代码",
            "   - 序列号：设备唯一序列号（必须唯一）",
            "   - 状态：active/inactive/maintenance",
            "   - 描述：设备详细描述",
            "",
            "2. 电机基本参数：",
            "   - 电机类型：三相感应电机/永磁同步电机等",
            f"   - 外径：{'mm' if is_metric else 'in'}（影响井筒适配性）",
            f"   - 长度：{'mm' if is_metric else 'in'}（影响安装空间）",
            f"   - 重量：{'kg' if is_metric else 'lbs'}（影响下井操作）",
            "   - 绝缘等级：B/F/H等（决定工作温度范围）",
            "   - 防护等级：IP54/IP55/IP68等（决定密封性能）",
            "",
            "3. ⚡ 频率参数字段（核心修复内容）：",
            "   🔑 这些字段解决了之前模板缺失电流电压信息的问题！",
            "",
            "   50Hz参数组：",
            f"   - 50Hz功率：{'kW' if is_metric else 'HP'}（额定功率）",
            "   - 50Hz电压：V（工作电压，如2300V、3300V、4160V）",
            "   - 50Hz电流：A（额定电流，用于电缆选型）⭐",
            "   - 50Hz转速：rpm（同步转速，通常2950rpm）",
            "",
            "   60Hz参数组：",
            f"   - 60Hz功率：{'kW' if is_metric else 'HP'}（额定功率）",
            "   - 60Hz电压：V（工作电压）",
            "   - 60Hz电流：A（额定电流，用于电缆选型）⭐",
            "   - 60Hz转速：rpm（同步转速，通常3540rpm）",
            "",
            "4. 🎯 电流电压信息的重要性：",
            "   ✅ 电流信息用于：",
            "      - 电缆截面积选型",
            "      - 变压器容量计算",
            "      - 保护器额定值设定",
            "      - 系统功率平衡分析",
            "",
            "   ✅ 电压信息用于：",
            "      - 电压等级匹配",
            "      - 绝缘要求确定",
            "      - 电网适配性评估",
            "",
            "5. 数据填写要求：",
            "   - 功率数值：不同频率下功率不同（60Hz通常比50Hz高20%）",
            "   - 电流数值：与功率和电压相关 I=P/(√3×U×cosφ×η)",
            "   - 转速数值：50Hz通常2950rpm，60Hz通常3540rpm",
            "   - 电压数值：常用2300V、3300V、4160V、6600V",
            "",
            "6. 单位转换说明：",
            f"   当前模板单位制：{'公制 (Metric)' if is_metric else '英制 (Imperial)'}",
            "   - 功率单位会自动转换为数据库标准单位",
            "   - 电流和电压单位统一使用A和V",
            "   - 转速统一使用rpm",
            "",
            "7. 常见参数范围：",
            "   - 功率范围：50-500HP (37-373kW)",
            "   - 电压范围：2300-6600V",
            "   - 电流范围：15-100A",
            "   - 转速：2950rpm(50Hz) / 3540rpm(60Hz)",
            "",
            "⚠️ 重要提醒：",
            "   1. ⭐ 电流和电压字段不能为空！",
            "   2. 序列号必须唯一",
            "   3. 同一电机的50Hz和60Hz参数必须都填写",
            "   4. 电流值直接影响电缆和保护器选型",
            "",
            "🔄 导入流程：",
            "   1. 按模板填写完整的电机参数",
            "   2. 确保50Hz和60Hz的电流电压数据完整",
            "   3. 保存并通过系统导入",
            "   4. 系统会自动创建频率参数记录",
            "   5. 导入后可在电机管理界面查看完整信息",
            "",
            "✅ 修复确认：",
            "   本模板已完全解决电流电压缺失问题！",
            "   所有电机选型所需的关键电气参数均已包含。"
        ]

        for i, note in enumerate(notes):
            cell = worksheet.cell(row=notes_start_row + i, column=1, value=note)
            if note.startswith("📋"):
                cell.font = Font(bold=True, size=14, color="2E7D32")
            elif note.startswith(("🔧", "1.", "2.", "3.", "4.", "5.", "6.", "7.", "⚠️", "🔄", "✅")):
                cell.font = Font(bold=True, color="1B5E20")
            elif "⭐" in note or "🔑" in note:
                cell.font = Font(bold=True, color="FF6F00")  # 重点标记用橙色
            else:
                cell.font = Font(size=10)
        
            # 合并单元格
            if note.strip():
                worksheet.merge_cells(
                    start_row=notes_start_row + i, 
                    start_column=1,
                end_row=notes_start_row + i,
                end_column=min(8, header_count)
            )

    def _build_motor_details(self, basic_record: Dict, is_metric: bool) -> Dict[str, Any]:
        """构建电机详细信息（匹配数据库字段）"""
        try:
            # 🔥 严格按照 device_motors 表字段构建
            motor_details = {
                'motor_type': str(basic_record.get('电机类型', '')),
                'outside_diameter': self._parse_float(basic_record.get('外径(mm)' if is_metric else '外径(in)')),
                'length': self._parse_float(basic_record.get('长度(mm)' if is_metric else '长度(in)')),
                'weight': self._parse_float(basic_record.get('重量(kg)' if is_metric else '重量(lbs)')),
                'insulation_class': str(basic_record.get('绝缘等级', '')),
                'protection_class': str(basic_record.get('防护等级', ''))
            }
        
            # 🔥 单位转换（如果需要）
            if not is_metric:
                # 将英制单位转换为公制（数据库标准）
                if motor_details['outside_diameter']:
                    motor_details['outside_diameter'] *= 25.4  # in -> mm
                if motor_details['length']:
                    motor_details['length'] *= 25.4  # in -> mm  
                if motor_details['weight']:
                    motor_details['weight'] *= 0.453592  # lbs -> kg
        
            # 🔥 注意：频率参数表 motor_frequency_params 需要单独处理
            # 这里暂时不处理频率参数，因为需要单独的表格结构
        
            logger.info(f"构建电机详细信息: {motor_details}")
            return motor_details
        
        except Exception as e:
            logger.error(f"构建电机详细信息失败: {str(e)}")
            raise

    def _generate_protector_template(self, workbook, is_metric: bool):
        """生成保护器导入模板（基于实际数据库结构）"""
        # 设置工作表
        if 'Sheet' in workbook.sheetnames:
            worksheet = workbook['Sheet']
        else:
            worksheet = workbook.create_sheet()

        worksheet.title = "保护器导入模板"

        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 🔥 根据实际数据库字段设置表头
        if is_metric:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（公制） - 对应 device_protectors 表字段
                "外径(mm)", "长度(mm)", "重量(kg)",
                "推力承载能力(kN)", "密封类型", "最高温度(°C)"
            ]
        
            example_data = [
                "Baker Hughes", "TandemSeal TS400", "TS400", "BH-TS400-001", "active", "高性能机械密封保护器",
                "114.3", "1524", "68.2",
                "89.0", "机械密封", "149"
            ]
        else:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（英制） - 对应 device_protectors 表字段
                "外径(in)", "长度(in)", "重量(lbs)",
                "推力承载能力(lbs)", "密封类型", "最高温度(°F)"
            ]
        
            example_data = [
                "Baker Hughes", "TandemSeal TS400", "TS400", "BH-TS400-001", "active", "高性能机械密封保护器",
                "4.5", "60", "150",
                "20000", "机械密封", "300"
            ]

        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 设置列宽
        column_widths = [15, 20, 10, 15, 8, 25, 12, 12, 12, 15, 12, 12]
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 填充示例数据
        for col, value in enumerate(example_data, 1):
            cell = worksheet.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # 添加更多示例数据
        additional_examples = []
        if is_metric:
            additional_examples = [
                ["Schlumberger", "REDA Max Protector", "MP450", "SLB-MP450-001", "active", "高温高压保护器",
                 "101.6", "1219", "54.5", "67.0", "组合密封", "177"],
                ["Weatherford", "Guardian Pro", "GP300", "WFT-GP300-001", "active", "标准机械密封保护器", 
                 "88.9", "914", "32.7", "45.0", "机械密封", "121"],
                ["Borets", "P-Series", "P450", "BOR-P450-001", "active", "经济型保护器",
                 "95.3", "1067", "41.3", "56.0", "迷宫密封", "135"]
            ]
        else:
            additional_examples = [
                ["Schlumberger", "REDA Max Protector", "MP450", "SLB-MP450-001", "active", "高温高压保护器",
                 "4.0", "48", "120", "15000", "组合密封", "350"],
                ["Weatherford", "Guardian Pro", "GP300", "WFT-GP300-001", "active", "标准机械密封保护器",
                 "3.5", "36", "72", "10000", "机械密封", "250"],
                ["Borets", "P-Series", "P450", "BOR-P450-001", "active", "经济型保护器",
                 "3.75", "42", "91", "12500", "迷宫密封", "275"]
            ]

        # 添加额外示例行
        for row_idx, example in enumerate(additional_examples, 3):
            for col, value in enumerate(example, 1):
                cell = worksheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 添加说明
        self._add_protector_template_notes(worksheet, is_metric, len(headers))

    def _add_protector_template_notes(self, worksheet, is_metric: bool, header_count: int):
        """添加保护器模板说明（基于实际数据库字段）"""
        notes_start_row = 7
    
        notes = [
            "📋 保护器导入模板说明：",
            "",
            "🔧 数据库字段对应：",
            "   本模板严格按照数据库表 device_protectors 的字段设计",
            "",
            "1. 基本信息字段：",
            "   - 制造商：设备制造商名称 (devices.manufacturer)",
            "   - 型号：完整设备型号 (devices.model)",
            "   - 系列：产品系列代码 (devices.series)",
            "   - 序列号：设备唯一序列号 (devices.serial_number)",
            "   - 状态：active/inactive/maintenance (devices.status)",
            "   - 描述：设备详细描述 (devices.description)",
            "",
            "2. 保护器专用字段 (device_protectors表)：",
            f"   - 外径：{'mm' if is_metric else 'in'} (outer_diameter)",
            f"   - 长度：{'mm' if is_metric else 'in'} (length)",
            f"   - 重量：{'kg' if is_metric else 'lbs'} (weight)",
            f"   - 推力承载能力：{'kN' if is_metric else 'lbs'} (thrust_capacity)",
            "   - 密封类型：机械密封/唇形密封/迷宫密封等 (seal_type)",
            f"   - 最高温度：{'°C' if is_metric else '°F'} (max_temperature)",
            "",
            "3. 数据要求：",
            "   - 所有数值字段支持小数",
            "   - 推力承载能力是关键参数，影响设备选型",
            "   - 密封类型影响适用工况",
            "   - 最高温度决定使用环境限制",
            "",
            "4. 单位说明：",
            f"   当前模板单位制：{'公制 (Metric)' if is_metric else '英制 (Imperial)'}",
            "   - 导入时系统会自动转换为数据库标准单位",
            "   - 推力承载能力：数据库存储为kN",
            "   - 温度：数据库存储为摄氏度",
            "   - 尺寸：数据库存储为mm",
            "",
            "5. 常见密封类型：",
            "   - 机械密封：适用于高压高转速",
            "   - 唇形密封：成本低，适用于标准工况",
            "   - 迷宫密封：无接触密封，寿命长",
            "   - 组合密封：多种密封形式组合",
            "",
            "⚠️ 注意事项：",
            "   1. 序列号必须唯一，重复将导致导入失败",
            "   2. 推力承载能力不能为空",
            "   3. 外径须满足套管尺寸要求",
            "   4. 密封类型请使用标准术语",
            "",
            "📞 技术支持：",
            "   如需了解更多字段含义或遇到导入问题，",
            "   请联系技术支持团队获取帮助。"
        ]

        for i, note in enumerate(notes):
            cell = worksheet.cell(row=notes_start_row + i, column=1, value=note)
            if note.startswith("📋"):
                cell.font = Font(bold=True, size=14, color="FF6F00")
            elif note.startswith(("🔧", "1.", "2.", "3.", "4.", "5.", "⚠️", "📞")):
                cell.font = Font(bold=True, color="D84315")
            else:
                cell.font = Font(size=10)
        
            # 合并单元格
            if note.strip():
                worksheet.merge_cells(
                    start_row=notes_start_row + i, 
                    start_column=1,
                    end_row=notes_start_row + i,
                    end_column=min(8, header_count)
                )

    def _build_protector_details(self, basic_record: Dict, is_metric: bool) -> Dict[str, Any]:
        """构建保护器详细信息（匹配数据库字段）"""
        try:
            # 🔥 严格按照 device_protectors 表字段构建
            protector_details = {
                'outer_diameter': self._parse_float(basic_record.get('外径(mm)' if is_metric else '外径(in)')),
                'length': self._parse_float(basic_record.get('长度(mm)' if is_metric else '长度(in)')),
                'weight': self._parse_float(basic_record.get('重量(kg)' if is_metric else '重量(lbs)')),
                'thrust_capacity': self._parse_float(basic_record.get('推力承载能力(kN)' if is_metric else '推力承载能力(lbs)')),
                'seal_type': str(basic_record.get('密封类型', '')),
                'max_temperature': self._parse_float(basic_record.get('最高温度(°C)' if is_metric else '最高温度(°F)'))
            }
        
            # 🔥 单位转换（如果需要）
            if not is_metric:
                # 将英制单位转换为公制（数据库标准）
                if protector_details['outer_diameter']:
                    protector_details['outer_diameter'] *= 25.4  # in -> mm
                if protector_details['length']:
                    protector_details['length'] *= 25.4  # in -> mm  
                if protector_details['weight']:
                    protector_details['weight'] *= 0.453592  # lbs -> kg
                if protector_details['thrust_capacity']:
                    protector_details['thrust_capacity'] *= 0.004448  # lbs -> kN
                if protector_details['max_temperature']:
                    protector_details['max_temperature'] = (protector_details['max_temperature'] - 32) * 5/9  # °F -> °C
        
            logger.info(f"构建保护器详细信息: {protector_details}")
            return protector_details
        
        except Exception as e:
            logger.error(f"构建保护器详细信息失败: {str(e)}")
            raise

    def _generate_separator_template(self, workbook, is_metric: bool):
        """生成分离器导入模板（基于实际数据库结构）"""
        # 设置工作表
        if 'Sheet' in workbook.sheetnames:
            worksheet = workbook['Sheet']
        else:
            worksheet = workbook.create_sheet()

        worksheet.title = "分离器导入模板"

        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # 🔥 根据实际数据库字段设置表头
        if is_metric:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（公制） - 对应 device_separators 表字段
                "外径(mm)", "长度(mm)", "重量(kg)",
                "分离效率(%)", "气体处理能力(m³/d)", "液体处理能力(m³/d)"
            ]
        
            example_data = [
                "Halliburton", "VORTEX-S500", "VORTEX", "HAL-VS500-001", "active", "高效旋流分离器",
                "152.4", "2438", "227",
                "95.5", "50000", "2000"
            ]
        else:
            headers = [
                "制造商", "型号", "系列", "序列号", "状态", "描述",
                # 基本参数（英制） - 对应 device_separators 表字段
                "外径(in)", "长度(in)", "重量(lbs)",
                "分离效率(%)", "气体处理能力(scf/d)", "液体处理能力(bbl/d)"
            ]
        
            example_data = [
                "Halliburton", "VORTEX-S500", "VORTEX", "HAL-VS500-001", "active", "高效旋流分离器",
                "6.0", "96", "500",
                "95.5", "1800000", "12600"
            ]

        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # 设置列宽
        column_widths = [15, 20, 10, 15, 8, 25, 12, 12, 12, 12, 18, 18]
        for col, width in enumerate(column_widths[:len(headers)], 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # 填充示例数据
        for col, value in enumerate(example_data, 1):
            cell = worksheet.cell(row=2, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # 添加更多示例数据
        additional_examples = []
        if is_metric:
            additional_examples = [
                ["Schlumberger", "HydroFrac Separator", "HFS300", "SLB-HFS300-001", "active", "高压气液分离器",
                 "127.0", "1829", "145", "92.0", "35000", "1500"],
                ["Weatherford", "Multi-Phase Separator", "MPS200", "WFT-MPS200-001", "active", "多相流分离器", 
                 "101.6", "1524", "89", "88.5", "25000", "1000"],
                ["Baker Hughes", "SUPER-SEP", "SS400", "BH-SS400-001", "active", "超级分离器",
                 "168.3", "3048", "310", "97.2", "75000", "3000"]
            ]
        else:
            additional_examples = [
                ["Schlumberger", "HydroFrac Separator", "HFS300", "SLB-HFS300-001", "active", "高压气液分离器",
                 "5.0", "72", "320", "92.0", "1260000", "9450"],
                ["Weatherford", "Multi-Phase Separator", "MPS200", "WFT-MPS200-001", "active", "多相流分离器",
                 "4.0", "60", "196", "88.5", "900000", "6300"],
                ["Baker Hughes", "SUPER-SEP", "SS400", "BH-SS400-001", "active", "超级分离器",
                 "6.625", "120", "683", "97.2", "2700000", "18900"]
            ]

        # 添加额外示例行
        for row_idx, example in enumerate(additional_examples, 3):
            for col, value in enumerate(example, 1):
                cell = worksheet.cell(row=row_idx, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # 添加说明
        self._add_separator_template_notes(worksheet, is_metric, len(headers))

    def _add_separator_template_notes(self, worksheet, is_metric: bool, header_count: int):
        """添加分离器模板说明（基于实际数据库字段）"""
        notes_start_row = 7
    
        notes = [
            "📋 分离器导入模板说明：",
            "",
            "🔧 数据库字段对应：",
            "   本模板严格按照数据库表 device_separators 的字段设计",
            "",
            "1. 基本信息字段：",
            "   - 制造商：设备制造商名称 (devices.manufacturer)",
            "   - 型号：完整设备型号 (devices.model)",
            "   - 系列：产品系列代码 (devices.series)",
            "   - 序列号：设备唯一序列号 (devices.serial_number)",
            "   - 状态：active/inactive/maintenance (devices.status)",
            "   - 描述：设备详细描述 (devices.description)",
            "",
            "2. 分离器专用字段 (device_separators表)：",
            f"   - 外径：{'mm' if is_metric else 'in'} (outer_diameter)",
            f"   - 长度：{'mm' if is_metric else 'in'} (length)",
            f"   - 重量：{'kg' if is_metric else 'lbs'} (weight)",
            "   - 分离效率：百分比 (separation_efficiency)",
            f"   - 气体处理能力：{'m³/d' if is_metric else 'scf/d'} (gas_handling_capacity)",
            f"   - 液体处理能力：{'m³/d' if is_metric else 'bbl/d'} (liquid_handling_capacity)",
            "",
            "3. 数据要求：",
            "   - 所有数值字段支持小数",
            "   - 分离效率范围：0-100%",
            "   - 处理能力是关键性能指标",
            "   - 外径须满足井筒尺寸要求",
            "",
            "4. 单位说明：",
            f"   当前模板单位制：{'公制 (Metric)' if is_metric else '英制 (Imperial)'}",
            "   - 导入时系统会自动转换为数据库标准单位",
            "   - 气体处理能力：数据库存储为m³/d",
            "   - 液体处理能力：数据库存储为m³/d",
            "   - 尺寸：数据库存储为mm",
            "",
            "5. 性能参数说明：",
            "   - 分离效率：气液分离的有效性指标",
            "   - 气体处理能力：单位时间可处理的气体体积",
            "   - 液体处理能力：单位时间可处理的液体体积",
            "   - 外径：影响安装空间和流通面积",
            "",
            "6. 常见分离器类型：",
            "   - 旋流分离器：利用离心力分离",
            "   - 重力分离器：利用密度差分离",
            "   - 多相流分离器：同时处理气液固三相",
            "   - 膜分离器：利用膜技术精确分离",
            "",
            "⚠️ 注意事项：",
            "   1. 序列号必须唯一，重复将导致导入失败",
            "   2. 分离效率不能超过100%",
            "   3. 处理能力须与实际工况匹配",
            "   4. 外径限制需考虑井筒空间",
            "",
            "📞 技术支持：",
            "   如需了解更多字段含义或遇到导入问题，",
            "   请联系技术支持团队获取帮助。"
        ]

        for i, note in enumerate(notes):
            cell = worksheet.cell(row=notes_start_row + i, column=1, value=note)
            if note.startswith("📋"):
                cell.font = Font(bold=True, size=14, color="9C27B0")
            elif note.startswith(("🔧", "1.", "2.", "3.", "4.", "5.", "6.", "⚠️", "📞")):
                cell.font = Font(bold=True, color="7B1FA2")
            else:
                cell.font = Font(size=10)
        
            # 合并单元格
            if note.strip():
                worksheet.merge_cells(
                    start_row=notes_start_row + i, 
                    start_column=1,
                    end_row=notes_start_row + i,
                    end_column=min(8, header_count)
                )

    def _build_separator_details(self, basic_record: Dict, is_metric: bool) -> Dict[str, Any]:
        """构建分离器详细信息（匹配数据库字段）"""
        try:
            # 🔥 严格按照 device_separators 表字段构建
            separator_details = {
                'outer_diameter': self._parse_float(basic_record.get('外径(mm)' if is_metric else '外径(in)')),
                'length': self._parse_float(basic_record.get('长度(mm)' if is_metric else '长度(in)')),
                'weight': self._parse_float(basic_record.get('重量(kg)' if is_metric else '重量(lbs)')),
                'separation_efficiency': self._parse_float(basic_record.get('分离效率(%)')),
                'gas_handling_capacity': self._parse_float(basic_record.get('气体处理能力(m³/d)' if is_metric else '气体处理能力(scf/d)')),
                'liquid_handling_capacity': self._parse_float(basic_record.get('液体处理能力(m³/d)' if is_metric else '液体处理能力(bbl/d)'))
            }
        
            # 🔥 单位转换（如果需要）
            if not is_metric:
                # 将英制单位转换为公制（数据库标准）
                if separator_details['outer_diameter']:
                    separator_details['outer_diameter'] *= 25.4  # in -> mm
                if separator_details['length']:
                    separator_details['length'] *= 25.4  # in -> mm  
                if separator_details['weight']:
                    separator_details['weight'] *= 0.453592  # lbs -> kg
                if separator_details['gas_handling_capacity']:
                    separator_details['gas_handling_capacity'] *= 0.0283168  # scf/d -> m³/d
                if separator_details['liquid_handling_capacity']:
                    separator_details['liquid_handling_capacity'] *= 0.158987  # bbl/d -> m³/d
        
            logger.info(f"构建分离器详细信息: {separator_details}")
            return separator_details
        
        except Exception as e:
            logger.error(f"构建分离器详细信息失败: {str(e)}")
            raise