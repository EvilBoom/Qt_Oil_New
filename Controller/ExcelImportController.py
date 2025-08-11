# Controller/ExcelImportController.py

from PySide6.QtCore import QObject, Signal, Slot, Property, QUrl
from typing import List, Dict, Optional, Any
import os
import json
import logging
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from DataManage.services.excel_import_service import ExcelImportService
from DataManage.services.database_service import DatabaseService

logger = logging.getLogger(__name__)

class ExcelImportController(QObject):
    """Excel导入控制器 - 处理井轨迹数据的Excel导入"""

    # 信号定义
    fileLoaded = Signal(str)                    # 文件路径
    columnsIdentified = Signal(dict)            # 列映射
    previewDataReady = Signal(list)             # 预览数据
    importProgress = Signal(int, int)           # 当前行, 总行数
    importCompleted = Signal(int, int)          # 井ID, 导入行数
    importFailed = Signal(str)                  # 错误信息
    sheetsLoaded = Signal(list)                 # 工作表列表
    validationCompleted = Signal(dict)          # 验证结果

    # 添加新的信号
    templateGenerated = Signal(str)             # 模板文件路径
    templateGenerationFailed = Signal(str)      # 模板生成失败

    def __init__(self):
        super().__init__()
        self._excel_service = ExcelImportService()
        self._db_service = DatabaseService()
        self._current_file_path = ""
        self._current_well_id = -1
        self._trajectory_data = []
        self._preview_data = []
        self._sheet_names = []
        self._import_summary = {}

    # ========== 属性定义 ==========
    @Property(str, notify=fileLoaded)
    def currentFilePath(self):
        """当前文件路径"""
        return self._current_file_path

    @Property(list, notify=previewDataReady)
    def previewData(self):
        """预览数据"""
        return self._preview_data

    @Property(list, notify=sheetsLoaded)
    def sheetNames(self):
        """工作表名称列表"""
        return self._sheet_names

    @Property(dict, notify=validationCompleted)
    def importSummary(self):
        """导入摘要"""
        return self._import_summary

    # ========== 文件操作 ==========

    @Slot(str)
    def loadExcelFile(self, file_url: str):
        """
        加载Excel文件

        Args:
            file_url: 文件URL（可能是file://开头）
        """
        try:
            # 处理文件URL
            if file_url.startswith('file:///'):
                file_path = file_url[8:]  # Windows
            elif file_url.startswith('file://'):
                file_path = file_url[7:]  # Unix/Mac
            else:
                file_path = file_url

            # 规范化路径
            file_path = os.path.normpath(file_path)

            if not os.path.exists(file_path):
                self.importFailed.emit(f"文件不存在: {file_path}")
                return

            # 检查文件扩展名
            if not file_path.lower().endswith(('.xls', '.xlsx')):
                self.importFailed.emit("请选择Excel文件(.xls或.xlsx)")
                return

            self._current_file_path = file_path

            # 获取工作表名称
            self._sheet_names = self._excel_service.get_sheet_names(file_path)
            self.sheetsLoaded.emit(self._sheet_names)

            # 加载第一个工作表
            if self._sheet_names:
                self.loadSheet(self._sheet_names[0])

            self.fileLoaded.emit(file_path)
            logger.info(f"Excel文件已加载: {file_path}")

        except Exception as e:
            error_msg = f"加载文件失败: {str(e)}"
            logger.error(error_msg)
            self.importFailed.emit(error_msg)

    @Slot(str)
    def loadSheet(self, sheet_name: str):
        """
        加载指定的工作表

        Args:
            sheet_name: 工作表名称
        """
        try:
            # 清空之前的数据
            self._excel_service.clear()

            # 读取Excel文件
            if not self._excel_service.read_excel_file(self._current_file_path, sheet_name):
                errors = self._excel_service.errors
                self.importFailed.emit("\n".join(errors) if errors else "读取Excel失败")
                return

            # 识别列
            column_mapping = self._excel_service.identify_columns()
            if not column_mapping:
                errors = self._excel_service.errors
                self.importFailed.emit("\n".join(errors) if errors else "无法识别必需的列")
                return

            self.columnsIdentified.emit(column_mapping)

            # 获取预览数据
            self._preview_data = self._excel_service.get_preview_data(20)
            self.previewDataReady.emit(self._preview_data)

            # 执行数据验证
            self.validateData()

        except Exception as e:
            error_msg = f"加载工作表失败: {str(e)}"
            logger.error(error_msg)
            self.importFailed.emit(error_msg)

    # ========== 数据验证 ==========

    @Slot()
    def validateData(self):
        """验证数据"""
        try:
            # 提取轨迹数据
            self._trajectory_data = self._excel_service.extract_trajectory_data()

            # 获取导入摘要
            self._import_summary = self._excel_service.get_import_summary()
            self._import_summary['data_count'] = len(self._trajectory_data)

            # 添加数据统计
            if self._trajectory_data:
                tvd_values = [d['tvd'] for d in self._trajectory_data]
                md_values = [d['md'] for d in self._trajectory_data]

                self._import_summary['statistics'] = {
                    'min_tvd': min(tvd_values),
                    'max_tvd': max(tvd_values),
                    'min_md': min(md_values),
                    'max_md': max(md_values),
                    'data_points': len(self._trajectory_data)
                }

            self.validationCompleted.emit(self._import_summary)

        except Exception as e:
            error_msg = f"数据验证失败: {str(e)}"
            logger.error(error_msg)
            self.importFailed.emit(error_msg)

    # ========== 数据导入 ==========

    @Slot(int)
    def importToWell_old(self, well_id: int):
        """
        导入数据到指定井

        Args:
            well_id: 井ID
        """
        if not self._trajectory_data:
            self.importFailed.emit("没有可导入的数据")
            return

        self._current_well_id = well_id

        try:
            # 准备导入数据
            import_data = []
            total_rows = len(self._trajectory_data)

            for idx, traj_data in enumerate(self._trajectory_data):
                # 发送进度
                self.importProgress.emit(idx + 1, total_rows)

                # 准备数据记录 - 修正字段映射
                record = {
                    'tvd': traj_data['tvd'],
                    'md': traj_data['md'],
                    'dls': traj_data.get('dls'),
                    'inclination': traj_data.get('inclination'),  # 修正：使用正确的字段名
                    'azimuth': traj_data.get('azimuth'),          # 修正：使用正确的字段名
                    'north_south': traj_data.get('north_south'),
                    'east_west': traj_data.get('east_west')
                }

                import_data.append(record)

            # 保存到数据库
            success = self._db_service.save_well_trajectories(well_id, import_data)

            if success:
                # 保存导入记录
                import_record = {
                    'well_id': well_id,
                    'file_name': os.path.basename(self._current_file_path),
                    'row_count': len(import_data),
                    'status': 'success',
                    'imported_by': 'current_user'  # TODO: 获取当前用户
                }

                # 添加统计信息到导入记录
                statistics = self._calculate_import_statistics(import_data)
                if statistics:
                    import_record['error_message'] = f"统计: {statistics}"

                if self._import_summary.get('warnings'):
                    warnings_text = '; '.join(self._import_summary['warnings'][:3])
                    import_record['error_message'] = (import_record.get('error_message', '') + 
                                                    f" 警告: {warnings_text}").strip()

                self._db_service.save_trajectory_import_record(import_record)

                self.importCompleted.emit(well_id, len(import_data))
                logger.info(f"成功导入{len(import_data)}条轨迹数据到井ID: {well_id}")
                logger.info(f"导入统计: {statistics}")
            else:
                self.importFailed.emit("保存数据到数据库失败")

        except Exception as e:
            error_msg = f"导入失败: {str(e)}"
            logger.error(error_msg)

            # 保存失败记录
            try:
                import_record = {
                    'well_id': well_id,
                    'file_name': os.path.basename(self._current_file_path),
                    'row_count': 0,
                    'status': 'failed',
                    'error_message': error_msg,
                    'imported_by': 'current_user'
                }
                self._db_service.save_trajectory_import_record(import_record)
            except:
                pass

            self.importFailed.emit(error_msg)
    

    @Slot('QVariant')  # 🔥 修改参数类型，接收QVariant对象
    def importToWell(self, import_params):
        """
        导入数据到指定井

        Args:
            import_params: 导入参数对象，包含wellId等信息
        """
        try:
            # 🔥 解析参数对象 - 处理QJSValue类型
            from PySide6.QtQml import QJSValue  


            if isinstance(import_params, QJSValue):
                # 🔥 将QJSValue转换为Python字典
                if import_params.isObject():
                    well_id = int(import_params.property('wellId').toNumber())
                    source_depth_unit = import_params.property('sourceDepthUnit').toString()
                    target_depth_unit = import_params.property('targetDepthUnit').toString()
                    is_metric = import_params.property('isMetric').toBool()
                    perform_unit_conversion = import_params.property('performUnitConversion').toBool()
                
                    logger.info(f"🔧 从QJSValue解析参数:")
                    logger.info(f"  - wellId: {well_id}")
                    logger.info(f"  - sourceDepthUnit: {source_depth_unit}")
                    logger.info(f"  - targetDepthUnit: {target_depth_unit}")
                    logger.info(f"  - isMetric: {is_metric}")
                    logger.info(f"  - performUnitConversion: {perform_unit_conversion}")
                else:
                    # 如果不是对象，可能是数字（井ID）
                    # 如果不是对象，可能是数字（井ID）
                    well_id = int(import_params.toNumber())
                    source_depth_unit = 'auto'
                    target_depth_unit = 'ft'
                    is_metric = False
                    perform_unit_conversion = False
                    logger.info(f"🔧 从QJSValue解析为数字: {well_id}")

            # 🔥 解析参数对象
            elif isinstance(import_params, dict):
                well_id = import_params.get('wellId', -1)
                source_depth_unit = import_params.get('sourceDepthUnit', 'auto')
                target_depth_unit = import_params.get('targetDepthUnit', 'ft')
                is_metric = import_params.get('isMetric', False)
                perform_unit_conversion = import_params.get('performUnitConversion', True)
                logger.info(f"🔧 从dict解析参数: {import_params}")

            elif isinstance(import_params, int):
                # 🔥 兼容旧版本调用方式
                well_id = import_params
                source_depth_unit = 'auto'
                target_depth_unit = 'ft'
                is_metric = False
                perform_unit_conversion = False
            else:
                logger.error(f"❌ 无效的导入参数类型: {type(import_params)}")
                self.importFailed.emit("无效的导入参数")
                return

            # 🔥 验证井ID
            if well_id <= 0:
                logger.error(f"❌ 井ID无效: {well_id}")
                self.importFailed.emit(f"井ID无效: {well_id}")
                return

            if not self._trajectory_data:
                self.importFailed.emit("没有可导入的数据")
                return

            self._current_well_id = well_id

            logger.info(f"🚀 开始导入数据:")
            logger.info(f"  - 井ID: {well_id}")
            logger.info(f"  - 源单位: {source_depth_unit}")
            logger.info(f"  - 目标单位: {target_depth_unit}")
            logger.info(f"  - 公制模式: {is_metric}")
            logger.info(f"  - 执行单位转换: {perform_unit_conversion}")
            logger.info(f"  - 数据条数: {len(self._trajectory_data)}")

            # 准备导入数据
            import_data = []
            total_rows = len(self._trajectory_data)

            for idx, traj_data in enumerate(self._trajectory_data):
                # 发送进度
                self.importProgress.emit(idx + 1, total_rows)

                # 🔥 准备数据记录 - 支持单位转换
                record = {
                    'tvd': self._convert_depth_value(traj_data['tvd'], source_depth_unit, target_depth_unit, perform_unit_conversion),
                    'md': self._convert_depth_value(traj_data['md'], source_depth_unit, target_depth_unit, perform_unit_conversion),
                    'dls': traj_data.get('dls'),
                    'inclination': traj_data.get('inclination'),
                    'azimuth': traj_data.get('azimuth'),
                    'north_south': traj_data.get('north_south'),
                    'east_west': traj_data.get('east_west')
                }

                import_data.append(record)

            # 保存到数据库
            success = self._db_service.save_well_trajectories(well_id, import_data)

            if success:
                # 保存导入记录
                import_record = {
                    'well_id': well_id,
                    'file_name': os.path.basename(self._current_file_path),
                    'row_count': len(import_data),
                    'status': 'success',
                    'imported_by': 'current_user',  # TODO: 获取当前用户
                    # 🔥 移除这些字段，因为模型不支持：
                    # 'source_unit': source_depth_unit,
                    # 'target_unit': target_depth_unit,
                    # 'unit_conversion': perform_unit_conversion
                }

                # 添加统计信息到导入记录
                statistics = self._calculate_import_statistics(import_data)
                if statistics:
                    import_record['error_message'] = f"统计: {statistics}"

                if self._import_summary.get('warnings'):
                    warnings_text = '; '.join(self._import_summary['warnings'][:3])
                    import_record['error_message'] = (import_record.get('error_message', '') + 
                                                    f" 警告: {warnings_text}").strip()

                self._db_service.save_trajectory_import_record(import_record)

                self.importCompleted.emit(well_id, len(import_data))
                logger.info(f"✅ 成功导入{len(import_data)}条轨迹数据到井ID: {well_id}")
                logger.info(f"导入统计: {statistics}")
            else:
                error_msg = "保存数据到数据库失败"
                logger.error(f"❌ {error_msg}")
                self.importFailed.emit(error_msg)

        except Exception as e:
            error_msg = f"导入失败: {str(e)}"
            logger.error(f"❌ {error_msg}")
            logger.error(f"参数详情: {import_params}")

            # 保存失败记录
            try:
                well_id = import_params.get('wellId', -1) if isinstance(import_params, dict) else import_params
                import_record = {
                    'well_id': well_id,
                    'file_name': os.path.basename(self._current_file_path) if self._current_file_path else 'unknown',
                    'row_count': 0,
                    'status': 'failed',
                    'error_message': error_msg,
                    'imported_by': 'current_user'
                }
                self._db_service.save_trajectory_import_record(import_record)
            except Exception as save_error:
                logger.error(f"保存失败记录时出错: {save_error}")

            self.importFailed.emit(error_msg)

    def _convert_depth_value(self, value, source_unit, target_unit, perform_conversion):
        """
        转换深度值单位
    
        Args:
            value: 原始值
            source_unit: 源单位
            target_unit: 目标单位
            perform_conversion: 是否执行转换
    
        Returns:
            转换后的值
        """
        if not perform_conversion or source_unit == target_unit or source_unit == 'auto':
            return value
    
        try:
            if source_unit == 'ft' and target_unit == 'm':
                return value * 0.3048  # 英尺转米
            elif source_unit == 'm' and target_unit == 'ft':
                return value / 0.3048  # 米转英尺
            else:
                return value
        except:
            return value

    
    def _calculate_import_statistics(self, import_data: List[Dict]) -> str:
        """计算导入统计信息"""
        if not import_data:
            return ""
        
        stats = []
        total = len(import_data)
        
        # 统计各字段的数据完整性
        field_counts = {}
        for record in import_data:
            for field in ['dls', 'inclination', 'azimuth', 'north_south', 'east_west']:
                if record.get(field) is not None:
                    field_counts[field] = field_counts.get(field, 0) + 1
        
        field_names = {
            'dls': '狗腿度',
            'inclination': '井斜角',
            'azimuth': '方位角',
            'north_south': '南北坐标',
            'east_west': '东西坐标'
        }
        
        for field, count in field_counts.items():
            percentage = (count / total) * 100
            stats.append(f"{field_names.get(field, field)}: {count}/{total}({percentage:.1f}%)")
        
        return '; '.join(stats)

    # ========== 辅助方法 ==========

    @Slot(result=dict)
    def getColumnMapping(self) -> dict:
        """获取列映射"""
        return self._excel_service.column_mapping.copy()

    @Slot(result=list)
    def getTrajectoryData(self) -> list:
        """获取轨迹数据"""
        return self._trajectory_data.copy()

    @Slot()
    def clearData(self):
        """清空数据"""
        self._excel_service.clear()
        self._current_file_path = ""
        self._trajectory_data = []
        self._preview_data = []
        self._sheet_names = []
        self._import_summary = {}

    @Slot(int, result=list)
    def getImportHistory(self, well_id: int) -> list:
        """
        获取导入历史

        Args:
            well_id: 井ID

        Returns:
            导入历史记录列表
        """
        try:
            # 这里需要在DatabaseService中添加相应的方法
            # 暂时返回空列表
            return []
        except Exception as e:
            logger.error(f"获取导入历史失败: {e}")
            return []

    @Slot(str, result=str)
    def generateDeviceTemplate(self, device_type: str) -> str:
        """
        生成设备导入模板
        
        Args:
            device_type: 设备类型 (pump, motor, protector, separator)
            
        Returns:
            模板文件路径
        """
        try:
            logger.info(f"开始生成{device_type}设备模板")
            
            # 创建工作簿
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            # 根据设备类型设置不同的模板
            if device_type == "pump":
                template_path = self._create_pump_template(worksheet)
            elif device_type == "motor":
                template_path = self._create_motor_template(worksheet)
            elif device_type == "protector":
                template_path = self._create_protector_template(worksheet)
            elif device_type == "separator":
                template_path = self._create_separator_template(worksheet)
            else:
                raise ValueError(f"不支持的设备类型: {device_type}")
            
            # 生成临时文件路径
            temp_dir = tempfile.gettempdir()
            filename = f"{device_type}_import_template.xlsx"
            file_path = os.path.join(temp_dir, filename)
            
            # 保存文件
            workbook.save(file_path)
            
            logger.info(f"模板生成成功: {file_path}")
            self.templateGenerated.emit(file_path)
            return file_path
            
        except Exception as e:
            error_msg = f"生成{device_type}模板失败: {str(e)}"
            logger.error(error_msg)
            self.templateGenerationFailed.emit(error_msg)
            return ""

    def _create_pump_template(self, worksheet):
        """创建泵设备导入模板"""
        worksheet.title = "泵设备导入模板"
        
        # 设置标题样式
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置边框
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 泵设备字段
        headers = [
            "制造商", "型号", "系列", "额定功率(HP)", "额定电压(V)", "额定频率(Hz)",
            "单级扬程(ft)", "单级功率(HP)", "最大级数", "最小流量(bbl/d)", "最大流量(bbl/d)",
            "效率(%)", "外径(in)", "轴径(in)", "重量(lbs)", "长度(in)", "状态", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            cell.border = border
            
        # 设置列宽
        column_widths = [12, 15, 10, 12, 12, 12, 12, 12, 10, 12, 12, 8, 10, 10, 10, 10, 8, 15]
        for col, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 添加示例数据
        example_data = [
            ["Centrilift", "GN4000", "GN", "250", "3300", "60", "12.5", "2.8", "150", "150", "2200", "78", "5.62", "1.5", "850", "25", "active", "高效率泵"],
            ["REDA", "DN1750", "DN", "180", "2300", "50", "10.2", "2.2", "120", "100", "1800", "75", "4.75", "1.3", "680", "22", "active", "标准泵"],
            ["Baker Hughes", "FLEX400", "FLEX", "300", "4160", "60", "15.0", "3.5", "180", "200", "2500", "80", "6.0", "1.8", "950", "28", "active", "大流量泵"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # 添加说明
        worksheet.cell(row=6, column=1, value="说明:")
        worksheet.cell(row=7, column=1, value="1. 请按照表头格式填写数据")
        worksheet.cell(row=8, column=1, value="2. 红色字段为必填项")
        worksheet.cell(row=9, column=1, value="3. 状态字段请填写: active 或 inactive")
        
        return "pump_template"

    def _create_motor_template(self, worksheet):
        """创建电机导入模板"""
        worksheet.title = "电机导入模板"
        
        # 设置样式
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 电机字段
        headers = [
            "制造商", "型号", "系列", "功率(HP)", "电压(V)", "频率(Hz)", "转速(RPM)",
            "电流(A)", "效率(%)", "功率因数", "绝缘等级", "防护等级", "外径(in)", 
            "长度(in)", "重量(lbs)", "温升(°C)", "状态", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            cell.border = border
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # 添加示例数据
        example_data = [
            ["Centrilift", "Electrospeed 562", "ES", "250", "3300", "60", "3600", "45", "92", "0.85", "F", "IP68", "5.62", "25", "450", "80", "active", "高效电机"],
            ["REDA", "Hotline HD", "HT", "180", "2300", "50", "3000", "38", "90", "0.82", "F", "IP68", "4.75", "22", "380", "75", "active", "耐高温"],
            ["Baker Hughes", "MaxForce", "MF", "300", "4160", "60", "3600", "42", "94", "0.88", "H", "IP68", "6.0", "28", "520", "85", "active", "大功率电机"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
        
        return "motor_template"

    def _create_protector_template(self, worksheet):
        """创建保护器导入模板"""
        worksheet.title = "保护器导入模板"
        
        # 设置样式
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 保护器字段
        headers = [
            "制造商", "型号", "系列", "额定推力(lbs)", "最大推力(lbs)", "转速(RPM)",
            "外径(in)", "长度(in)", "重量(lbs)", "油容量(gal)", "轴承类型", 
            "密封类型", "工作温度(°F)", "状态", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            cell.border = border
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # 添加示例数据
        example_data = [
            ["Centrilift", "P562", "P", "15000", "18000", "3600", "5.62", "12", "85", "2.5", "球轴承", "机械密封", "350", "active", "标准保护器"],
            ["REDA", "Sentinel", "S", "12000", "15000", "3000", "4.75", "10", "65", "2.0", "滚子轴承", "机械密封", "300", "active", "紧凑型"],
            ["Baker Hughes", "Guardian", "G", "20000", "25000", "3600", "6.0", "15", "120", "3.0", "混合轴承", "双密封", "400", "active", "重载保护器"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
        
        return "protector_template"

    def _create_separator_template(self, worksheet):
        """创建分离器导入模板"""
        worksheet.title = "分离器导入模板"
        
        # 设置样式
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 分离器字段
        headers = [
            "制造商", "型号", "系列", "分离效率(%)", "最大流量(bbl/d)", "最大气液比",
            "外径(in)", "长度(in)", "重量(lbs)", "入口压力(psi)", "出口压力(psi)",
            "工作温度(°F)", "材质", "状态", "备注"
        ]
        
        # 设置表头
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            cell.border = border
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
        
        # 添加示例数据
        example_data = [
            ["Centrilift", "GSEP562", "GSEP", "95", "2000", "500", "5.62", "8", "45", "2000", "1950", "300", "不锈钢", "active", "高效分离器"],
            ["REDA", "Vortex", "VX", "92", "1800", "400", "4.75", "7", "38", "1800", "1750", "280", "碳钢", "active", "旋流分离器"],
            ["Baker Hughes", "HydroCyclone", "HC", "98", "2500", "600", "6.0", "10", "55", "2500", "2450", "350", "合金钢", "active", "水力旋流器"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = border
        
        return "separator_template"

    @Slot(str)
    def downloadTemplate(self, device_type: str):
        """
        下载模板文件
        
        Args:
            device_type: 设备类型
        """
        try:
            template_path = self.generateDeviceTemplate(device_type)
            if template_path:
                # 打开文件管理器并选中文件
                import subprocess
                import platform
                
                if platform.system() == "Windows":
                    subprocess.run(f'explorer /select,"{template_path}"', shell=True)
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", "-R", template_path])
                else:  # Linux
                    subprocess.run(["xdg-open", os.path.dirname(template_path)])
                
                logger.info(f"模板文件已生成并打开: {template_path}")
            
        except Exception as e:
            error_msg = f"下载模板失败: {str(e)}"
            logger.error(error_msg)
            self.templateGenerationFailed.emit(error_msg)
